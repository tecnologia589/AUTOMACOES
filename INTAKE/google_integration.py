import os
import sys
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/documents',
    'https://www.googleapis.com/auth/spreadsheets'
]

def autenticar_google():
    """
    Autentica com OAuth2 (conta pessoal do usuario) se oauth_credentials.json existir.
    Caso contrario, usa Service Account como fallback.
    """
    # Buscar credenciais na pasta config/ (raiz do projeto)
    project_root = os.path.join(os.path.dirname(__file__), '..')
    config_dir = os.path.join(project_root, 'config')
    oauth_creds_path = os.getenv("GOOGLE_OAUTH_CREDENTIALS", os.path.join(config_dir, "oauth_credentials.json"))
    token_path = os.path.join(config_dir, "token.json")
    sa_creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", os.path.join(config_dir, "credentials.json"))

    creds = None

    # Tenta OAuth2 primeiro (conta pessoal - usa cota do usuario)
    if os.path.exists(oauth_creds_path):
        if os.path.exists(token_path):
            creds = Credentials.from_authorized_user_file(token_path, SCOPES)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                print("Renovando token de acesso...")
                creds.refresh(Request())
            else:
                print("Abrindo navegador para login com sua conta Google...")
                print("(isso so precisa ser feito uma vez)")
                flow = InstalledAppFlow.from_client_secrets_file(oauth_creds_path, SCOPES)
                creds = flow.run_local_server(port=0)

            with open(token_path, 'w') as token:
                token.write(creds.to_json())

        print("Autenticado com sua conta Google pessoal!")

    # Fallback: Service Account
    elif os.path.exists(sa_creds_path):
        print("AVISO: Usando Service Account (cota limitada).")
        print("Para usar sua conta pessoal, configure oauth_credentials.json")
        creds = service_account.Credentials.from_service_account_file(
            sa_creds_path, scopes=SCOPES)
    else:
        print(f"ERRO: Nenhum arquivo de credenciais encontrado.")
        print("Coloque 'oauth_credentials.json' ou 'credentials.json' na pasta do projeto.")
        sys.exit(1)

    try:
        drive_service = build('drive', 'v3', credentials=creds)
        docs_service = build('docs', 'v1', credentials=creds)
        return drive_service, docs_service
    except Exception as e:
        print(f"ERRO DE AUTENTICACAO: {str(e)}")
        sys.exit(1)

def duplicar_template(drive_service, template_id: str, nome_cliente: str, data: str, pasta_destino_id: str = None) -> str:
    """
    Duplica o arquivo de template e o renomeia, mantendo na mesma pasta original
    ou enviando para uma pasta destino específica.
    """
    print("Gerando cópia do Template Ficha no Google Drive...")

    try:
        nome_arquivo = f"{nome_cliente} - Ficha Cliente - {data}"
        body = {
            'name': nome_arquivo,
            'mimeType': 'application/vnd.google-apps.document'
        }

        if pasta_destino_id:
            body['parents'] = [pasta_destino_id]
        else:
            arquivo_origem = drive_service.files().get(fileId=template_id, fields='parents').execute()
            if 'parents' in arquivo_origem:
                body['parents'] = arquivo_origem['parents']

        copia = drive_service.files().copy(
            fileId=template_id, body=body).execute()

        return copia.get('id')

    except HttpError as error:
        print(f"ERRO DE API DRIVE: Falha ao duplicar o arquivo.")
        print("Você compartilhou a pasta destino (ou o template original) com o email robô?")
        print(f"Detalhes técnicos: {error}")
        sys.exit(1)


def criar_pasta_cliente(drive_service, nome_cliente: str) -> str:
    """
    Cria a pasta do cliente dentro de RECLAMANTE com as 3 subpastas padrao:
    - ATOS INTERNOS
    - DOCUMENTOS DO CLIENTE
    - PASTA DO CLIENTE
    Retorna o ID da pasta principal criada.
    """
    pasta_reclamante_id = os.getenv("GOOGLE_PASTA_RECLAMANTE", "")
    if not pasta_reclamante_id:
        print("ERRO: GOOGLE_PASTA_RECLAMANTE nao configurado no .env (ID da pasta-mae no Drive).")
        return None

    print(f"Criando pasta do cliente: {nome_cliente}...")

    try:
        # Criar pasta principal do cliente
        pasta_cliente = drive_service.files().create(body={
            'name': nome_cliente.upper(),
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [pasta_reclamante_id]
        }, fields='id, webViewLink').execute()

        pasta_id = pasta_cliente['id']
        pasta_link = pasta_cliente.get('webViewLink', '')

        # Criar as 3 subpastas
        subpastas = ['ATOS INTERNOS', 'DOCUMENTOS DO CLIENTE', 'PASTA DO CLIENTE']
        for sub in subpastas:
            drive_service.files().create(body={
                'name': sub,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [pasta_id]
            }).execute()
            print(f"   Subpasta criada: {sub}")

        print(f"Pasta do cliente criada com sucesso!")
        print(f"Link: {pasta_link}")
        return pasta_id

    except HttpError as error:
        print(f"ERRO ao criar pasta do cliente: {error}")
        return None


def buscar_proximo_contrato(drive_service, nome_cliente="", empresa="", area="TRABALHISTA"):
    """
    Busca o proximo numero sequencial de contrato na planilha de controle
    e registra o novo contrato na proxima linha.
    """
    import openpyxl
    import io
    from datetime import datetime
    from googleapiclient.http import MediaInMemoryUpload

    planilha_id = os.getenv("GOOGLE_PLANILHA_CONTRATOS", "")
    if not planilha_id:
        print("   Aviso: GOOGLE_PLANILHA_CONTRATOS nao configurado; pulando numeracao de contrato.")
        return ""

    ano_atual = str(datetime.now().year)
    aba_contratos = os.getenv("GOOGLE_ABA_CONTRATOS", f"Contratos {ano_atual}")

    try:
        content = drive_service.files().get_media(fileId=planilha_id).execute()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb[aba_contratos]

        # Encontrar ultimo numero e ultima linha preenchida
        sufixo_ano = f'/{ano_atual}'
        ultimo = 0
        ultima_linha = 4
        for row in range(5, ws.max_row + 1):
            val = str(ws.cell(row=row, column=8).value or '')
            if sufixo_ano in val:
                try:
                    num = int(val.split('/')[0])
                    if num > ultimo:
                        ultimo = num
                        ultima_linha = row
                except:
                    pass

        proximo = ultimo + 1
        nova_linha = ultima_linha + 1
        numero_contrato = f"{proximo}/{ano_atual}"

        # Preencher nova linha no padrao
        ws.cell(row=nova_linha, column=4, value=nome_cliente.upper())  # D: Nome
        ws.cell(row=nova_linha, column=6, value=datetime.now())  # F: Data cadastro
        ws.cell(row=nova_linha, column=8, value=numero_contrato)  # H: CT.
        ws.cell(row=nova_linha, column=9, value=empresa.upper())  # I: Parte contraria
        ws.cell(row=nova_linha, column=10, value=area)  # J: Area

        # Salvar e fazer upload de volta
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        media = MediaInMemoryUpload(
            buffer.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        drive_service.files().update(
            fileId=planilha_id,
            media_body=media
        ).execute()

        print(f"   Planilha atualizada: linha {nova_linha} - {nome_cliente.upper()} - {numero_contrato}")
        return f"CT. {numero_contrato}"

    except Exception as e:
        print(f"   Aviso: Nao conseguiu atualizar planilha: {e}")
        return ""


def gerar_documentos_cliente(drive_service, docs_service, nome_cliente: str, dados: dict):
    """
    Gera Contrato, Procuracao e Declaracao de Hipossuficiencia
    a partir dos templates, preenchendo com os dados do cliente.

    IDs de template e pastas de destino vem do .env (cliente fornece os DELE):
      GOOGLE_TEMPLATE_CONTRATO / GOOGLE_PASTA_CONTRATO
      GOOGLE_TEMPLATE_PROCURACAO / GOOGLE_PASTA_PROCURACAO
      GOOGLE_TEMPLATE_DECLARACAO / GOOGLE_PASTA_DECLARACAO
    """
    from datetime import datetime

    templates = {
        'Contrato de Honorarios': {
            'template_id': os.getenv("GOOGLE_TEMPLATE_CONTRATO", ""),
            'pasta_id': os.getenv("GOOGLE_PASTA_CONTRATO", ""),
        },
        'Procuracao': {
            'template_id': os.getenv("GOOGLE_TEMPLATE_PROCURACAO", ""),
            'pasta_id': os.getenv("GOOGLE_PASTA_PROCURACAO", ""),
        },
        'Declaracao de Hipossuficiencia': {
            'template_id': os.getenv("GOOGLE_TEMPLATE_DECLARACAO", ""),
            'pasta_id': os.getenv("GOOGLE_PASTA_DECLARACAO", ""),
        },
    }

    hoje = datetime.now().strftime('%d de %B de %Y').replace(
        'January', 'Janeiro').replace('February', 'Fevereiro').replace(
        'March', 'Marco').replace('April', 'Abril').replace(
        'May', 'Maio').replace('June', 'Junho').replace(
        'July', 'Julho').replace('August', 'Agosto').replace(
        'September', 'Setembro').replace('October', 'Outubro').replace(
        'November', 'Novembro').replace('December', 'Dezembro')

    cidade_foro = os.getenv("ESCRITORIO_CIDADE", "Cáceres/MT")
    data_local = f"{cidade_foro}, {hoje}"

    # Buscar numero sequencial do contrato
    empresa = dados.get('qual_empresa', dados.get('empresa', ''))
    numero_contrato = buscar_proximo_contrato(drive_service, nome_cliente, empresa)
    print(f"   Numero do contrato: {numero_contrato}")

    # Mapeamento das chaves dos templates para os dados extraidos
    mapa_chaves = {
        '{{Nome do cliente}}': dados.get('nome', ''),
        '{{Nome completo}}': dados.get('nome', ''),
        '{{CPF do cliente}}': dados.get('cpf', ''),
        '{{CPF}}': dados.get('cpf', ''),
        '{{RG do cliente}}': dados.get('rg', ''),
        '{{RG}}': dados.get('rg', ''),
        '{{Nacionalidade do cliente}}': dados.get('nacionalidade', ''),
        '{{Nacionalidade}}': dados.get('nacionalidade', ''),
        '{{Estado civil do cliente}}': dados.get('estado_civil', ''),
        '{{Estado Civil}}': dados.get('estado_civil', ''),
        '{{Profissão do cliente}}': dados.get('profissao', ''),
        '{{Profissão}}': dados.get('profissao', ''),
        '{{ENDEREÇO DO CLIENTE}}': dados.get('endereco', ''),
        '{{Endereço}}': dados.get('endereco', ''),
        '{{Bairro}}': dados.get('bairro', ''),
        '{{Cidade Estado}}': dados.get('cidade_estado', ''),
        '{{CEP}}': dados.get('cep', ''),
        '{{telefone do cliente}}': dados.get('telefone', ''),
        '{{e-mail do cliente}}': dados.get('email', ''),
        '{{NOME DA AÇÃO}}': dados.get('tipo_acao', 'Reclamatoria Trabalhista'),
        '{{NOME DA EMPRESA}}': dados.get('qual_empresa', ''),
        '{{nome da empresa}}': dados.get('qual_empresa', ''),
        '{{Data e local de hoje}}': data_local,
        '{{data e cidade}}': data_local,
        '{{ Número sequencial de contrato  EX: CT. N°.182/2024}}': numero_contrato,
    }

    docs_gerados = []

    for doc_nome, config in templates.items():
        if not config['template_id'] or not config['pasta_id']:
            print(f"   Aviso: template/pasta de '{doc_nome}' nao configurado no .env; pulando.")
            continue

        print(f"   Gerando: {nome_cliente} - {doc_nome}...")

        try:
            nome_arquivo = f"{nome_cliente} - {doc_nome}"
            copia = drive_service.files().copy(
                fileId=config['template_id'],
                body={
                    'name': nome_arquivo,
                    'mimeType': 'application/vnd.google-apps.document',
                    'parents': [config['pasta_id']]
                }
            ).execute()

            doc_id = copia['id']

            # Preencher as chaves
            requests = []
            for chave, valor in mapa_chaves.items():
                if valor:
                    requests.append({
                        'replaceAllText': {
                            'containsText': {'text': chave, 'matchCase': False},
                            'replaceText': valor
                        }
                    })

            if requests:
                docs_service.documents().batchUpdate(
                    documentId=doc_id, body={'requests': requests}).execute()

            docs_gerados.append({'nome': nome_arquivo, 'id': doc_id})
            print(f"   Pronto: {nome_arquivo}")

        except HttpError as error:
            print(f"   ERRO ao gerar {doc_nome}: {error}")

    return docs_gerados


def preencher_documento(docs_service, document_id: str, dados: dict):
    """
    Edita um documento Google via Requests de BatchUpdate substituindo os placeholders
    {{CHAVE}} com os valores recém extraídos pela IA no dicionário.
    """
    print("Inserindo Fatos e Detalhes no novo documento gerado...")

    requests = []

    # Mapeamento do dicionário do Claude com as Tags reais
    # Note que se a IA botou "FATOS" e "NOME_CLIENTE", vamos buscar "{{FATOS}}" no doc.
    for chave, valor in dados.items():
        if not isinstance(valor, str):
            # Se vier uma lista/dicionário maluco, converte
            valor = str(valor)

        requests.append({
            'replaceAllText': {
                'containsText': {
                    'text': f"{{{{{chave}}}}}",
                    'matchCase': True
                },
                'replaceText': valor
            }
        })

    try:
        resultado = docs_service.documents().batchUpdate(
            documentId=document_id, body={'requests': requests}).execute()
        print("Documento finalizado com Sucesso!!")

    except HttpError as error:
        print(f"ERRO DE API DOCS: Houve um erro ao formatar o texto do arquivo.")
        print(f"Detalhes: {error}")
        sys.exit(1)
