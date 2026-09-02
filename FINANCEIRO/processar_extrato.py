"""
Processa extrato Asaas e gera planilha de fechamento de um mes.
Cria a aba do mes + uma aba por comissionado cadastrado em config.

A classificacao de comissoes vem 100% de config/regras_financeiras.py
(COMISSOES), que comeca VAZIO.
>>> TODO (onboarding): cadastre os comissionados/percentuais/exclusoes do
>>> escritorio em config/regras_financeiras.py antes de usar em producao.

Uso:
  py processar_extrato.py 03/2026
  py processar_extrato.py 03/2026 --sem-upload
"""
import sys, io, os, re, argparse
from datetime import datetime, timedelta

import requests

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'INTEGRACOES'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'config'))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', 'config', '.env'))

from google_integration import autenticar_google
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from googleapiclient.http import MediaInMemoryUpload
import httplib2
import google_auth_httplib2

import regras_financeiras as regras

# IDs de pasta do Drive - VAZIOS por padrao (cliente cria os dele).
PASTA_FINANCEIRO_API = os.getenv('DRIVE_PASTA_FINANCEIRO_ID', '')

ASAAS_API_TOKEN = os.getenv('ASAAS_API_TOKEN', '')
ASAAS_BASE_URL = os.getenv('ASAAS_BASE_URL', 'https://api.asaas.com/v3')

MESES_NOME = {
    '01': 'JANEIRO', '02': 'FEVEREIRO', '03': 'MARCO', '04': 'ABRIL',
    '05': 'MAIO', '06': 'JUNHO', '07': 'JULHO', '08': 'AGOSTO',
    '09': 'SETEMBRO', '10': 'OUTUBRO', '11': 'NOVEMBRO', '12': 'DEZEMBRO'
}
MESES_ABREV = {
    '01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR',
    '05': 'MAI', '06': 'JUN', '07': 'JUL', '08': 'AGO',
    '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'
}


def baixar_extrato_asaas(start_date, finish_date):
    """Baixa extrato financeiro direto da API Asaas com paginacao."""
    headers = {'access_token': ASAAS_API_TOKEN}
    todas_transacoes = []
    offset = 0
    limit = 100

    while True:
        params = {
            'startDate': start_date,
            'finishDate': finish_date,
            'offset': offset,
            'limit': limit
        }
        resp = requests.get(f'{ASAAS_BASE_URL}/financialTransactions',
                            headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        todas_transacoes.extend(data['data'])
        print(f'  Baixados {len(todas_transacoes)}/{data["totalCount"]} registros...')
        if not data['hasMore']:
            break
        offset += limit

    return todas_transacoes


# Mapeamento de tipos da API -> tipo legivel
TIPOS_CREDITO = {'PAYMENT_RECEIVED', 'RECEIVABLE_ANTICIPATION', 'CHARGEBACK_REVERSAL',
                 'TRANSFER_RECEIVED', 'REFUND_REVERSAL'}
TIPOS_DEBITO = {'TRANSFER', 'PAYMENT_FEE', 'PAYMENT_REFUND', 'CHARGEBACK',
                'RECEIVABLE_ANTICIPATION_FEE', 'BILL_PAYMENT', 'PAYMENT_FEE_REFUND'}


def extrair_nome_cliente(desc):
    """Extrai nome base do cliente da descricao, removendo detalhes."""
    desc = str(desc)
    if 'fatura nr.' in desc:
        desc = re.split(r'fatura nr\.\s*\d+\s*', desc)[-1]
    nome = re.split(r'\s*\[', desc)[0].strip()
    nome = re.sub(r'[_\s]*(ATUAL)$', '', nome).strip()
    nome = re.sub(r'\.\s*Nota fiscal nr\.\s*\d+', '', nome).strip()
    nome = re.sub(r'^recebida\s*-\s*ATRASADOS\s*CONTA\s*\w+\s*-\s*', '', nome).strip()
    return nome.upper()


# === ESTILOS ===
header_font = Font(name='Montserrat', bold=True, size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font_white = Font(name='Montserrat', bold=True, size=11, color='FFFFFF')
normal_font = Font(name='Montserrat', size=10)
money_fmt = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
credito_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
debito_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
comissao_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


def processar(competencia, sem_upload=False):
    mes, ano = competencia.split('/')
    start_date = f'{ano}-{mes}-01'
    if int(mes) == 12:
        end_dt = datetime(int(ano) + 1, 1, 1) - timedelta(days=1)
    else:
        end_dt = datetime(int(ano), int(mes) + 1, 1) - timedelta(days=1)
    end_date = end_dt.strftime('%Y-%m-%d')
    nome_mes = MESES_NOME[mes]
    abrev = MESES_ABREV[mes]
    ano_curto = ano[2:]

    drive, _ = autenticar_google()

    # ========================================================
    # PASSO 1: BAIXAR EXTRATO VIA API ASAAS
    # ========================================================
    print('Baixando extrato via API Asaas...')
    transacoes_api = baixar_extrato_asaas(start_date, end_date)
    print(f'  Total: {len(transacoes_api)} transacoes')

    # Converter para formato interno + classificar comissao (via config)
    dados = []
    for t in transacoes_api:
        valor = t['value']
        tipo_lanc = 'Crédito' if valor > 0 else 'Débito'
        descricao = t.get('description') or ''
        parceiro = ''
        if tipo_lanc == 'Crédito' and 'Cobrança recebida' in descricao:
            parceiro = regras.classificar_comissao(descricao) or ''

        dados.append({
            'data': t.get('date', ''),
            'tipo_trans': t.get('type', ''),
            'descricao': descricao,
            'valor': abs(valor),
            'saldo': t.get('balance', 0),
            'fatura': t.get('paymentId', ''),
            'tipo_lanc': tipo_lanc,
            'parceiro': parceiro
        })

    creditos = [d for d in dados if d['tipo_lanc'] == 'Crédito']
    debitos = [d for d in dados if d['tipo_lanc'] == 'Débito']
    cobr_recebidas = [d for d in creditos if 'Cobrança recebida' in d['descricao']]

    # Mostrar classificacao
    print('\n--- Classificacao dos creditos (via config) ---')
    for d in cobr_recebidas:
        nome = extrair_nome_cliente(d['descricao'])
        rotulo = regras.COMISSOES.get(d['parceiro'], {}).get('rotulo', 'ESCRITORIO') if d['parceiro'] else 'ESCRITORIO'
        print(f'  [{rotulo:^12}] R$ {d["valor"]:>10,.2f} | {nome}')

    # === CRIAR PLANILHA NOVA ===
    print('\nCriando planilha...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ========================================================
    # ABA 1: MES
    # ========================================================
    aba_mes = f'{nome_mes}_{ano_curto}'
    print(f'Criando aba {aba_mes}...')
    ws = wb.create_sheet(aba_mes, 0)

    ws.merge_cells('A1:I1')
    ws['A1'] = f'FECHAMENTO {nome_mes}/{ano}'
    ws['A1'].font = Font(name='Montserrat', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Resumo
    ws['A3'] = 'RESUMO DO MES'
    ws['A3'].font = Font(name='Montserrat', bold=True, size=12)

    ws['A4'] = 'Total Creditos:'
    ws['B4'] = sum(d['valor'] for d in creditos)
    ws['B4'].number_format = money_fmt
    ws['A5'] = 'Total Debitos:'
    ws['B5'] = sum(d['valor'] for d in debitos)
    ws['B5'].number_format = money_fmt
    ws['A6'] = 'Cobrancas Recebidas:'
    ws['B6'] = sum(d['valor'] for d in cobr_recebidas)
    ws['B6'].number_format = money_fmt

    # Resumo por comissionado (dinamico)
    linha_resumo = 4
    ws.cell(row=linha_resumo, column=4, value='ESCRITORIO:')
    ws.cell(row=linha_resumo, column=5,
            value=sum(d['valor'] for d in cobr_recebidas if not d['parceiro'])).number_format = money_fmt
    for chave, regra in regras.COMISSOES.items():
        linha_resumo += 1
        ws.cell(row=linha_resumo, column=4, value=f"{regra.get('rotulo', chave)}:")
        ws.cell(row=linha_resumo, column=5,
                value=sum(d['valor'] for d in cobr_recebidas if d['parceiro'] == chave)).number_format = money_fmt

    for r in range(4, linha_resumo + 1):
        ws.cell(row=r, column=1).font = Font(name='Montserrat', bold=True, size=10)
        ws.cell(row=r, column=4).font = Font(name='Montserrat', bold=True, size=10)

    # Headers
    headers = ['Data', 'Tipo Transacao', 'Descricao', 'Valor', 'Saldo', 'Fatura',
               'Tipo', 'Comissionado', 'Comissao']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Dados
    row_n = 10
    for d in dados:
        rotulo = regras.COMISSOES.get(d['parceiro'], {}).get('rotulo', '') if d['parceiro'] else ''
        ws.cell(row=row_n, column=1, value=d['data']).font = normal_font
        ws.cell(row=row_n, column=2, value=d['tipo_trans']).font = normal_font
        ws.cell(row=row_n, column=3, value=d['descricao']).font = normal_font
        c4 = ws.cell(row=row_n, column=4, value=d['valor'])
        c4.font = normal_font
        c4.number_format = money_fmt
        c5 = ws.cell(row=row_n, column=5, value=d['saldo'])
        c5.font = normal_font
        c5.number_format = money_fmt
        ws.cell(row=row_n, column=6, value=d['fatura']).font = normal_font
        ws.cell(row=row_n, column=7, value=d['tipo_lanc']).font = normal_font
        ws.cell(row=row_n, column=8, value=rotulo).font = normal_font

        # Cores por tipo
        if d['tipo_lanc'] == 'Crédito':
            for c in range(1, 8):
                ws.cell(row=row_n, column=c).fill = credito_fill
        elif d['tipo_lanc'] == 'Débito':
            for c in range(1, 8):
                ws.cell(row=row_n, column=c).fill = debito_fill

        if d['parceiro']:
            ws.cell(row=row_n, column=8).fill = comissao_fill

        # Comissao (via regra configurada)
        nome_cli = extrair_nome_cliente(d['descricao'])
        if d['parceiro'] and 'Cobrança recebida' in d['descricao']:
            regra = regras.COMISSOES.get(d['parceiro'], {})
            pct = regra.get('percentual', 0)
            if pct and regras.comissionado_recebe(d['parceiro'], nome_cli):
                ws.cell(row=row_n, column=9, value=d['valor'] * pct).number_format = money_fmt

        for c in range(1, 10):
            ws.cell(row=row_n, column=c).border = thin_border
        row_n += 1

    # Totais
    row_t = row_n + 1
    ws.cell(row=row_t, column=3, value='TOTAL COMISSOES:').font = header_font
    total_comissoes = sum(ws.cell(row=r, column=9).value or 0 for r in range(10, row_n))
    ws.cell(row=row_t, column=9, value=total_comissoes).number_format = money_fmt
    ws.cell(row=row_t, column=9).font = header_font

    for col, w in zip('ABCDEFGHI', [12, 20, 65, 14, 14, 14, 10, 18, 16]):
        ws.column_dimensions[col].width = w

    print(f'  -> {row_n - 10} linhas de dados')

    # ========================================================
    # ABAS POR COMISSIONADO (uma por chave em COMISSOES)
    # ========================================================
    idx_aba = 1
    totais = {}
    for chave, regra in regras.COMISSOES.items():
        pct = regra.get('percentual', 0)
        rotulo = regra.get('rotulo', chave)
        processos = [d for d in cobr_recebidas
                     if d['parceiro'] == chave
                     and regras.comissionado_recebe(chave, extrair_nome_cliente(d['descricao']))]
        excluidos = [d for d in cobr_recebidas
                     if d['parceiro'] == chave
                     and not regras.comissionado_recebe(chave, extrair_nome_cliente(d['descricao']))]

        print(f'Criando aba {rotulo} {abrev}.{ano_curto} ({len(processos)} processos)...')
        if excluidos:
            print(f'  Excluidos de {rotulo}:')
            for d in excluidos:
                print(f'    {extrair_nome_cliente(d["descricao"])}: R$ {d["valor"]:,.2f}')

        ws_c = wb.create_sheet(f'{rotulo[:20]} {abrev}.{ano_curto}', idx_aba)
        idx_aba += 1
        ws_c.merge_cells('A1:D1')
        ws_c['A1'] = f'EXTRATO {nome_mes} - COMISSAO {rotulo.upper()}'
        ws_c['A1'].font = Font(name='Montserrat', bold=True, size=13)
        ws_c['B3'] = f'Periodo de 01/{mes}/{ano} ate {end_date[8:]}/{mes}/{ano}'
        ws_c['B3'].font = Font(name='Montserrat', bold=True, size=10)

        for col, h in enumerate(['Data', 'Descricao', 'FATURAMENTO', f'COMISSAO [{pct*100:.0f}%]'], 1):
            cell = ws_c.cell(row=4, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        c_row = 5
        for d in processos:
            ws_c.cell(row=c_row, column=1, value=d['data']).font = normal_font
            ws_c.cell(row=c_row, column=2, value=d['descricao']).font = normal_font
            c3 = ws_c.cell(row=c_row, column=3, value=d['valor'])
            c3.font = normal_font
            c3.number_format = money_fmt
            ws_c.cell(row=c_row, column=4, value=d['valor'] * pct).number_format = money_fmt
            for c in range(1, 5):
                ws_c.cell(row=c_row, column=c).border = thin_border
            c_row += 1

        c_row += 1
        total = sum(d['valor'] * pct for d in processos)
        ws_c.cell(row=c_row, column=2, value='TOTAIS:').font = header_font
        ws_c.cell(row=c_row, column=3, value=sum(d['valor'] for d in processos)).number_format = money_fmt
        ws_c.cell(row=c_row, column=3).font = header_font
        ws_c.cell(row=c_row, column=4, value=total).number_format = money_fmt
        ws_c.cell(row=c_row, column=4).font = header_font

        for col, w in zip('ABCD', [12, 70, 16, 24]):
            ws_c.column_dimensions[col].width = w

        totais[chave] = {'rotulo': rotulo, 'total': total, 'qtd': len(processos)}

    # ========================================================
    # UPLOAD
    # ========================================================
    if sem_upload or not PASTA_FINANCEIRO_API:
        local_path = os.path.join(os.path.dirname(__file__), f'FECHAMENTO {nome_mes} {ano}.xlsx')
        wb.save(local_path)
        if not PASTA_FINANCEIRO_API and not sem_upload:
            print('\nAVISO: DRIVE_PASTA_FINANCEIRO_ID nao configurado - salvando local.')
        print(f'  Salvo local: {local_path}')
    else:
        print('\nSalvando e enviando pro Drive...')
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.read()
        media = MediaInMemoryUpload(
            file_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_metadata = {
            'name': f'FECHAMENTO {nome_mes} {ano}',
            'parents': [PASTA_FINANCEIRO_API],
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        http_upload = google_auth_httplib2.AuthorizedHttp(
            drive._http.credentials,
            http=httplib2.Http(timeout=300)
        )
        import time
        for tentativa in range(3):
            try:
                arquivo = drive.files().create(
                    body=file_metadata, media_body=media, fields='id, webViewLink'
                ).execute(http=http_upload)
                print(f'  Arquivo criado: {arquivo.get("webViewLink")}')
                break
            except Exception as e:
                print(f'  Tentativa {tentativa+1} falhou: {e}')
                if tentativa == 2:
                    raise
                time.sleep(5)
                media = MediaInMemoryUpload(
                    file_bytes,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    # Resumo final
    print('\n' + '=' * 50)
    print('CONCLUIDO!')
    print('=' * 50)
    print(f'Aba {aba_mes}: {row_n - 10} linhas')
    for chave, info in totais.items():
        print(f'Aba {info["rotulo"]}: {info["qtd"]} processos | Comissao: R$ {info["total"]:,.2f}')
    if not totais:
        print('Nenhuma comissao cadastrada (COMISSOES vazio em config/regras_financeiras.py).')


def main():
    parser = argparse.ArgumentParser(description='Processa extrato Asaas e gera planilha de fechamento')
    parser.add_argument('competencia', help='Competencia MM/YYYY (ex: 03/2026)')
    parser.add_argument('--sem-upload', action='store_true', help='Salva local em vez de enviar pro Drive')
    args = parser.parse_args()
    processar(args.competencia, sem_upload=args.sem_upload)


if __name__ == '__main__':
    main()
