"""
Preenche a planilha de Resultado Financeiro anual - Corbelino Advogados Associados.

Puxa receitas e despesas do ADVBOX (fonte da verdade) para a competencia
informada, calcula os indicadores e grava na linha do mes correspondente.

NAO contem numeros hardcoded: tudo vem da API do ADVBOX do escritorio.
A categorizacao (comissoes / taxas / sazonais / distribuicao) usa as
palavras-chave configuraveis de config/regras_financeiras.py.

>>> TODO (onboarding): configure DRIVE_PLANILHA_RESULTADO_ID no config/.env
>>> (planilha anual do escritorio) e ajuste as palavras-chave de categoria em
>>> config/regras_financeiras.py conforme o plano de contas do escritorio.

Uso:
  py preencher_resultado.py 03/2026
"""
import sys, io, os, argparse
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'INTEGRACOES'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'config'))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', 'config', '.env'))

from google_integration import autenticar_google
import openpyxl
from googleapiclient.http import MediaInMemoryUpload

from advbox_integration import listar_transacoes
import regras_financeiras as regras

# ID da planilha anual de resultado - VAZIO por padrao (cliente cria a dele).
PLANILHA_ID = os.getenv('DRIVE_PLANILHA_RESULTADO_ID', '')

MESES_NOME = {
    '01': 'JANEIRO', '02': 'FEVEREIRO', '03': 'MARCO', '04': 'ABRIL',
    '05': 'MAIO', '06': 'JUNHO', '07': 'JULHO', '08': 'AGOSTO',
    '09': 'SETEMBRO', '10': 'OUTUBRO', '11': 'NOVEMBRO', '12': 'DEZEMBRO'
}


def coletar_advbox(competencia):
    """Puxa receitas/despesas do ADVBOX e agrupa por categoria."""
    mes, ano = competencia.split('/')
    start_date = f'{ano}-{mes}-01'
    if int(mes) == 12:
        end_dt = datetime(int(ano) + 1, 1, 1) - timedelta(days=1)
    else:
        end_dt = datetime(int(ano), int(mes) + 1, 1) - timedelta(days=1)
    end_date = end_dt.strftime('%Y-%m-%d')

    trans = listar_transacoes({'date_due_start': start_date, 'date_due_end': end_date})

    receitas = [t for t in trans
                if t.get('entry_type') == 'income'
                and not regras.excluir_do_faturamento(t.get('name') or '')]
    despesas = [t for t in trans if t.get('entry_type') == 'expense']

    faturamento_bruto = sum(t.get('amount', 0) for t in receitas)
    total_despesas_advbox = sum(t.get('amount', 0) for t in despesas)

    cats = defaultdict(float)
    for t in despesas:
        cats[(t.get('category') or 'SEM CATEGORIA').upper()] += t.get('amount', 0)

    distribuicao = sum(v for k, v in cats.items() if regras.eh_distribuicao(k))
    comissoes = sum(v for k, v in cats.items()
                    if any(p in k for p in regras.PALAVRAS_COMISSOES))
    taxas = sum(v for k, v in cats.items()
                if any(p in k for p in regras.PALAVRAS_TAXAS))
    sazonais = sum(v for k, v in cats.items()
                   if any(p in k for p in regras.PALAVRAS_SAZONAIS))

    return {
        'faturamento_bruto': faturamento_bruto,
        'total_despesas_advbox': total_despesas_advbox,
        'distribuicao': distribuicao,
        'comissoes': comissoes,
        'taxas': taxas,
        'sazonais': sazonais,
    }


def preencher(competencia):
    mes, ano = competencia.split('/')
    nome_mes = MESES_NOME[mes]

    dados = coletar_advbox(competencia)
    faturamento_bruto = dados['faturamento_bruto']
    total_despesas_advbox = dados['total_despesas_advbox']
    distribuicao_lucros = dados['distribuicao']
    comissoes_parceiros = dados['comissoes']
    despesas_sazonais = dados['sazonais']
    taxas_impostos = dados['taxas']

    # Distribuicao de lucros NAO e despesa operacional
    despesas_operacionais = total_despesas_advbox - distribuicao_lucros

    honorarios = faturamento_bruto
    faturamento_sem_parceiro = honorarios - comissoes_parceiros
    lucro_liquido = honorarios - despesas_operacionais
    margem_lucro = (lucro_liquido / honorarios * 100) if honorarios > 0 else 0
    perc_despesa_receita = (despesas_operacionais / honorarios * 100) if honorarios > 0 else 0
    perc_distribuicao = (distribuicao_lucros / honorarios * 100) if honorarios > 0 else 0
    provisao = lucro_liquido * regras.PERCENTUAL_PROVISAO_LUCRO if lucro_liquido > 0 else 0

    # PRINT
    print()
    print('=' * 60)
    print(f'  ANALISE FINANCEIRA {nome_mes}/{ano}')
    print('=' * 60)
    print(f'  Faturamento (ADVBOX):         R$ {faturamento_bruto:>12,.2f}')
    print(f'  Fat. s/ Parceiro:            R$ {faturamento_sem_parceiro:>12,.2f}')
    print(f'  ---')
    print(f'  Despesas Sazonais:           R$ {despesas_sazonais:>12,.2f}')
    print(f'  Comissao Parceiros:          R$ {comissoes_parceiros:>12,.2f}')
    print(f'  Taxas e Impostos:            R$ {taxas_impostos:>12,.2f}')
    print(f'  Distribuicao Lucros:         R$ {distribuicao_lucros:>12,.2f}')
    print(f'  ---')
    print(f'  DESPESAS TOTAIS (ADVBOX):    R$ {total_despesas_advbox:>12,.2f}')
    print(f'  DESP. OPERACIONAIS:          R$ {despesas_operacionais:>12,.2f}')
    print(f'  ---')
    print(f'  LUCRO LIQUIDO:               R$ {lucro_liquido:>12,.2f}')
    print(f'  Margem de lucro:             {margem_lucro:.1f}%')
    print(f'  % Despesa/Receita:           {perc_despesa_receita:.1f}%')
    print(f'  % Distribuicao/Faturamento:  {perc_distribuicao:.1f}%')
    print(f'  {regras.ROTULO_PROVISAO_LUCRO}: R$ {provisao:>12,.2f}')

    if not PLANILHA_ID:
        print('\nAVISO: DRIVE_PLANILHA_RESULTADO_ID nao configurado - nada gravado.')
        print('>>> TODO (onboarding): defina o ID da planilha anual no config/.env.')
        return

    # Baixar planilha
    print('\nBaixando planilha de resultado...')
    drive, _ = autenticar_google()
    content = drive.files().get_media(fileId=PLANILHA_ID).execute()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb[ano] if ano in wb.sheetnames else wb.active

    # ========================================================
    # PREENCHER PLANILHA - linha do mes (Jan=4, Fev=5, ...)
    # Colunas:
    # A=Nome Mes, B=Fat.Bruto, C=Hon.c/inadimplencia, D=hon.s/inadimplencia
    # E=Fat.s/parceiro, F=Emprestimos, G=Desp.SAZONAIS, H=Comissao Parceiros
    # I=Taxas e Impostos, J=Distribuicao, K=Despesas Totais
    # L=Lucro Liquido, M=margem de lucro, N=Fluxo de caixa
    # ========================================================
    linha = int(mes) + 3

    ws.cell(row=linha, column=1, value=nome_mes)
    ws.cell(row=linha, column=2, value=round(faturamento_bruto, 2))
    ws.cell(row=linha, column=3, value=round(honorarios, 2))
    ws.cell(row=linha, column=4, value=round(honorarios, 2))
    ws.cell(row=linha, column=5, value=round(faturamento_sem_parceiro, 2))
    # col 6 = emprestimos (nao tem)
    ws.cell(row=linha, column=7, value=round(despesas_sazonais, 2))
    ws.cell(row=linha, column=8, value=round(comissoes_parceiros, 2))
    ws.cell(row=linha, column=9, value=round(taxas_impostos, 2))
    ws.cell(row=linha, column=10, value=round(distribuicao_lucros, 2))
    ws.cell(row=linha, column=11, value=round(total_despesas_advbox, 2))
    ws.cell(row=linha, column=12, value=round(lucro_liquido, 2))
    ws.cell(row=linha, column=13, value=round(margem_lucro, 2))

    # Salvar e upload
    print('Salvando e enviando pro Drive...')
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.read()

    media = MediaInMemoryUpload(
        file_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    drive.files().update(
        fileId=PLANILHA_ID,
        media_body=media
    ).execute()

    print(f'PLANILHA ATUALIZADA! Linha {linha} ({nome_mes})')
    print(f'https://docs.google.com/spreadsheets/d/{PLANILHA_ID}/edit')


def main():
    parser = argparse.ArgumentParser(description='Preenche planilha de resultado financeiro anual')
    parser.add_argument('competencia', help='Competencia MM/YYYY (ex: 03/2026)')
    args = parser.parse_args()
    preencher(args.competencia)


if __name__ == '__main__':
    main()
