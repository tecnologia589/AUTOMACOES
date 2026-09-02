"""
Conciliacao Financeira: Asaas x ADVBOX - Corbelino Advogados Associados
Compara transacoes do Asaas (gateway de pagamento) com lancamentos do ADVBOX
(gestao juridica). Identifica divergencias, lancamentos faltantes e gera
relatorio de conciliacao.

Uso: py conciliar_financeiro.py [MM/YYYY]
  Ex: py conciliar_financeiro.py 04/2026
  Sem argumento: usa mes atual
"""
import sys, io, os, re, argparse
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'INTEGRACOES'))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', 'config', '.env'))

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from google_integration import autenticar_google
from googleapiclient.http import MediaInMemoryUpload
import httplib2
import google_auth_httplib2

from advbox_integration import (
    listar_transacoes, carregar_settings, buscar_categoria,
    buscar_banco, buscar_centro_custo, criar_transacao,
    buscar_cliente, resumo_financeiro, listar_settings_tipo
)

ASAAS_API_TOKEN = os.getenv('ASAAS_API_TOKEN', '')
ASAAS_BASE_URL = os.getenv('ASAAS_BASE_URL', 'https://api.asaas.com/v3')
# ID de pasta do Drive - VAZIO por padrao (cliente cria a dele).
PASTA_FINANCEIRO_API = os.getenv('DRIVE_PASTA_FINANCEIRO_ID', '')


# ============================================================
# ASAAS - Baixar extrato
# ============================================================

def baixar_extrato_asaas(start_date, finish_date):
    """Baixa extrato financeiro da API Asaas com paginacao."""
    headers = {'access_token': ASAAS_API_TOKEN}
    todas = []
    offset = 0
    limit = 100
    while True:
        params = {
            'startDate': start_date, 'finishDate': finish_date,
            'offset': offset, 'limit': limit
        }
        resp = requests.get(f'{ASAAS_BASE_URL}/financialTransactions',
                            headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        todas.extend(data['data'])
        print(f'  Asaas: {len(todas)}/{data["totalCount"]} registros...')
        if not data['hasMore']:
            break
        offset += limit
    return todas


TIPOS_CREDITO = {'PAYMENT_RECEIVED', 'RECEIVABLE_ANTICIPATION', 'CHARGEBACK_REVERSAL',
                 'TRANSFER_RECEIVED', 'REFUND_REVERSAL'}
TIPOS_DEBITO = {'TRANSFER', 'PAYMENT_FEE', 'PAYMENT_REFUND', 'CHARGEBACK',
                'RECEIVABLE_ANTICIPATION_FEE', 'BILL_PAYMENT', 'PAYMENT_FEE_REFUND'}


def extrair_nome_cliente(desc):
    """Extrai nome base do cliente da descricao Asaas."""
    desc = str(desc)
    if 'fatura nr.' in desc:
        desc = re.split(r'fatura nr\.\s*\d+\s*', desc)[-1]
    nome = re.split(r'\s*\[', desc)[0].strip()
    nome = re.sub(r'[_\s]*(ATUAL)$', '', nome).strip()
    nome = re.sub(r'\.\s*Nota fiscal nr\.\s*\d+', '', nome).strip()
    nome = re.sub(r'^recebida\s*-\s*ATRASADOS\s*CONTA\s*\w+\s*-\s*', '', nome).strip()
    return nome.upper()


# ============================================================
# NORMALIZACAO PARA MATCHING
# ============================================================

def normalizar_valor(valor):
    """Normaliza valor para float com 2 casas."""
    if isinstance(valor, str):
        valor = valor.replace('.', '').replace(',', '.')
    return round(float(valor), 2)


def normalizar_nome(nome):
    """Normaliza nome para comparacao."""
    if not nome:
        return ''
    nome = str(nome).upper().strip()
    nome = re.sub(r'[_\s]*(ATUAL)$', '', nome)
    nome = re.sub(r'\s+', ' ', nome).strip()
    return nome


def nomes_similares(nome1, nome2):
    """Verifica se dois nomes referem ao mesmo cliente (match flexivel)."""
    n1 = normalizar_nome(nome1)
    n2 = normalizar_nome(nome2)
    if not n1 or not n2:
        return False
    if n1 == n2:
        return True
    if n1 in n2 or n2 in n1:
        return True
    partes1 = n1.split()
    partes2 = n2.split()
    if len(partes1) >= 2 and len(partes2) >= 2:
        if partes1[0] == partes2[0] and partes1[-1] == partes2[-1]:
            return True
    return False


# ============================================================
# CONCILIACAO PRINCIPAL
# ============================================================

def conciliar(competencia_mmyyyy):
    """
    Executa conciliacao completa entre Asaas e ADVBOX.

    1. Baixa transacoes do Asaas (extrato do mes)
    2. Baixa transacoes do ADVBOX (lancamentos do mes)
    3. Cruza por valor + nome do cliente + data proxima
    4. Identifica: conciliados, so Asaas, so ADVBOX, divergencias de valor
    5. Gera planilha com resultado

    Retorna dict com resultado da conciliacao.
    """
    mes, ano = competencia_mmyyyy.split('/')
    start_date = f'{ano}-{mes}-01'
    if int(mes) == 12:
        end_date = f'{int(ano)+1}-01-01'
    else:
        end_date = f'{ano}-{int(mes)+1:02d}-01'
    end_date_dt = datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=1)
    end_date = end_date_dt.strftime('%Y-%m-%d')

    meses_nome = {
        '01': 'JANEIRO', '02': 'FEVEREIRO', '03': 'MARCO', '04': 'ABRIL',
        '05': 'MAIO', '06': 'JUNHO', '07': 'JULHO', '08': 'AGOSTO',
        '09': 'SETEMBRO', '10': 'OUTUBRO', '11': 'NOVEMBRO', '12': 'DEZEMBRO'
    }
    nome_mes = meses_nome.get(mes, mes)

    print(f'\n{"="*60}')
    print(f'  CONCILIACAO FINANCEIRA - {nome_mes}/{ano}')
    print(f'  Periodo: {start_date} a {end_date}')
    print(f'{"="*60}\n')

    # ----- PASSO 1: ASAAS -----
    print('1. Baixando extrato Asaas...')
    transacoes_asaas = baixar_extrato_asaas(start_date, end_date)
    print(f'   Total Asaas: {len(transacoes_asaas)} transacoes\n')

    asaas_items = []
    for t in transacoes_asaas:
        valor = t['value']
        descricao = t.get('description') or ''
        tipo = 'receita' if valor > 0 else 'despesa'
        nome_cli = extrair_nome_cliente(descricao) if 'Cobrança recebida' in descricao else ''
        eh_cobranca = 'Cobrança recebida' in descricao
        asaas_items.append({
            'data': t.get('date', ''),
            'tipo': tipo,
            'descricao': descricao,
            'valor': abs(valor),
            'nome_cliente': nome_cli,
            'fatura': t.get('paymentId', ''),
            'tipo_asaas': t.get('type', ''),
            'eh_cobranca': eh_cobranca,
            'conciliado': False,
            'advbox_match': None,
        })

    # ----- PASSO 2: ADVBOX -----
    print('2. Baixando lancamentos ADVBOX...')
    advbox_filtros = {
        'date_due_start': start_date,
        'date_due_end': end_date,
    }
    transacoes_advbox = listar_transacoes(advbox_filtros)
    print(f'   Total ADVBOX: {len(transacoes_advbox)} lancamentos\n')

    # Tambem buscar por competencia
    advbox_filtros_comp = {
        'competence_start': competencia_mmyyyy,
        'competence_end': competencia_mmyyyy,
    }
    transacoes_advbox_comp = listar_transacoes(advbox_filtros_comp)

    advbox_ids = {t['id'] for t in transacoes_advbox}
    for t in transacoes_advbox_comp:
        if t['id'] not in advbox_ids:
            transacoes_advbox.append(t)
            advbox_ids.add(t['id'])

    print(f'   Total ADVBOX (com competencia): {len(transacoes_advbox)} lancamentos\n')

    advbox_items = []
    for t in transacoes_advbox:
        advbox_items.append({
            'id': t.get('id'),
            'data_venc': t.get('date_due', ''),
            'data_pgto': t.get('date_payment', ''),
            'tipo': 'receita' if t.get('entry_type') == 'income' else 'despesa',
            'valor': normalizar_valor(t.get('amount', 0)),
            'descricao': t.get('description', '') or '',
            'categoria': t.get('category', '') or '',
            'nome_cliente': normalizar_nome(t.get('name', '') or ''),
            'cpf': t.get('identification', '') or '',
            'processo_id': t.get('lawsuit_id'),
            'num_processo': t.get('process_number', '') or '',
            'responsavel': t.get('responsible', '') or '',
            'banco': t.get('debit_bank', '') or '',
            'centro_custo': t.get('cost_center', '') or '',
            'conciliado': False,
            'asaas_match': None,
        })

    # ----- PASSO 3: MATCHING -----
    print('3. Conciliando transacoes...')

    conciliados = []
    # Match por valor exato + nome similar + data proxima (ate 5 dias)
    for a in asaas_items:
        if a['conciliado'] or not a['eh_cobranca']:
            continue
        for b in advbox_items:
            if b['conciliado'] or b['tipo'] != 'receita':
                continue
            if abs(a['valor'] - b['valor']) > 0.01:
                continue
            if a['nome_cliente'] and b['nome_cliente']:
                if not nomes_similares(a['nome_cliente'], b['nome_cliente']):
                    continue
            try:
                data_a = datetime.strptime(a['data'], '%Y-%m-%d')
                data_b = datetime.strptime(b['data_venc'], '%Y-%m-%d') if b['data_venc'] else None
                data_b2 = datetime.strptime(b['data_pgto'], '%Y-%m-%d') if b['data_pgto'] else None
                match_data = False
                if data_b and abs((data_a - data_b).days) <= 5:
                    match_data = True
                if data_b2 and abs((data_a - data_b2).days) <= 5:
                    match_data = True
                if not match_data and data_b:
                    if data_a.month == data_b.month and data_a.year == data_b.year:
                        match_data = True
            except (ValueError, TypeError):
                match_data = True

            if match_data:
                a['conciliado'] = True
                a['advbox_match'] = b['id']
                b['conciliado'] = True
                b['asaas_match'] = a['fatura']
                conciliados.append({
                    'asaas': a,
                    'advbox': b,
                    'tipo_match': 'valor+nome+data',
                })
                break

    # Segunda passada: match so por valor exato (cobrancas sem nome)
    for a in asaas_items:
        if a['conciliado'] or not a['eh_cobranca']:
            continue
        candidatos = [b for b in advbox_items
                      if not b['conciliado'] and b['tipo'] == 'receita'
                      and abs(a['valor'] - b['valor']) <= 0.01]
        if len(candidatos) == 1:
            b = candidatos[0]
            a['conciliado'] = True
            a['advbox_match'] = b['id']
            b['conciliado'] = True
            b['asaas_match'] = a['fatura']
            conciliados.append({
                'asaas': a,
                'advbox': b,
                'tipo_match': 'valor_unico',
            })

    # Match de despesas/debitos (taxas Asaas vs despesas ADVBOX)
    for a in asaas_items:
        if a['conciliado'] or a['tipo'] != 'despesa':
            continue
        for b in advbox_items:
            if b['conciliado'] or b['tipo'] != 'despesa':
                continue
            if abs(a['valor'] - b['valor']) <= 0.01:
                a['conciliado'] = True
                a['advbox_match'] = b['id']
                b['conciliado'] = True
                b['asaas_match'] = a.get('fatura', '')
                conciliados.append({
                    'asaas': a,
                    'advbox': b,
                    'tipo_match': 'despesa_valor',
                })
                break

    # Separar resultados
    so_asaas = [a for a in asaas_items if not a['conciliado']]
    so_advbox = [b for b in advbox_items if not b['conciliado']]
    so_asaas_cobrancas = [a for a in so_asaas if a['eh_cobranca']]
    so_asaas_taxas = [a for a in so_asaas if a['tipo'] == 'despesa']
    so_asaas_outros = [a for a in so_asaas if not a['eh_cobranca'] and a['tipo'] != 'despesa']

    # ----- RESUMO -----
    total_asaas_receita = sum(a['valor'] for a in asaas_items if a['tipo'] == 'receita')
    total_asaas_despesa = sum(a['valor'] for a in asaas_items if a['tipo'] == 'despesa')
    total_advbox_receita = sum(b['valor'] for b in advbox_items if b['tipo'] == 'receita')
    total_advbox_despesa = sum(b['valor'] for b in advbox_items if b['tipo'] == 'despesa')

    resultado = {
        'competencia': competencia_mmyyyy,
        'nome_mes': nome_mes,
        'ano': ano,
        'conciliados': conciliados,
        'so_asaas_cobrancas': so_asaas_cobrancas,
        'so_asaas_taxas': so_asaas_taxas,
        'so_asaas_outros': so_asaas_outros,
        'so_advbox': so_advbox,
        'total_asaas_receita': total_asaas_receita,
        'total_asaas_despesa': total_asaas_despesa,
        'total_advbox_receita': total_advbox_receita,
        'total_advbox_despesa': total_advbox_despesa,
        'asaas_items': asaas_items,
        'advbox_items': advbox_items,
    }

    print(f'\n--- RESULTADO DA CONCILIACAO ---')
    print(f'  Conciliados:          {len(conciliados)}')
    print(f'  So no Asaas (cobr.):  {len(so_asaas_cobrancas)}')
    print(f'  So no Asaas (taxas):  {len(so_asaas_taxas)}')
    print(f'  So no Asaas (outros): {len(so_asaas_outros)}')
    print(f'  So no ADVBOX:         {len(so_advbox)}')
    print(f'\n  Asaas  -> Receitas: R$ {total_asaas_receita:,.2f} | Despesas: R$ {total_asaas_despesa:,.2f}')
    print(f'  ADVBOX -> Receitas: R$ {total_advbox_receita:,.2f} | Despesas: R$ {total_advbox_despesa:,.2f}')
    diff_receita = total_asaas_receita - total_advbox_receita
    diff_despesa = total_asaas_despesa - total_advbox_despesa
    print(f'  Diferenca Receitas: R$ {diff_receita:,.2f}')
    print(f'  Diferenca Despesas: R$ {diff_despesa:,.2f}')

    return resultado


# ============================================================
# GERAR PLANILHA DE CONCILIACAO
# ============================================================

def gerar_planilha(resultado):
    """Gera planilha Excel com resultado da conciliacao."""
    nome_mes = resultado['nome_mes']
    ano = resultado['ano']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font = Font(name='Montserrat', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    normal = Font(name='Montserrat', size=10)
    bold = Font(name='Montserrat', bold=True, size=10)
    money = '#,##0.00'
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    blue_fill = PatternFill(start_color='DAE8FC', end_color='DAE8FC', fill_type='solid')

    def hdr_row(ws, row, headers, widths=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border
        if widths:
            for i, w in enumerate(widths):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w

    def data_cell(ws, row, col, val, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = normal
        c.border = border
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
        return c

    # ============ ABA 1: RESUMO ============
    ws = wb.create_sheet('RESUMO', 0)
    ws.merge_cells('A1:F1')
    ws['A1'] = f'CONCILIACAO {nome_mes}/{ano} - ASAAS x ADVBOX'
    ws['A1'].font = Font(name='Montserrat', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = 'TOTAIS'
    ws['A3'].font = Font(name='Montserrat', bold=True, size=12)

    labels = [
        ('Receitas Asaas:', resultado['total_asaas_receita']),
        ('Receitas ADVBOX:', resultado['total_advbox_receita']),
        ('Diferenca Receitas:', resultado['total_asaas_receita'] - resultado['total_advbox_receita']),
        ('', ''),
        ('Despesas Asaas:', resultado['total_asaas_despesa']),
        ('Despesas ADVBOX:', resultado['total_advbox_despesa']),
        ('Diferenca Despesas:', resultado['total_asaas_despesa'] - resultado['total_advbox_despesa']),
    ]
    for i, (label, val) in enumerate(labels):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = bold
        if val != '':
            c = ws.cell(row=r, column=2, value=val)
            c.number_format = money
            c.font = bold
            if 'Diferenca' in label and val != 0:
                c.fill = red_fill if val > 0 else yellow_fill

    r = 13
    stats = [
        ('Conciliados:', len(resultado['conciliados']), green_fill),
        ('So no Asaas (cobrancas):', len(resultado['so_asaas_cobrancas']), red_fill),
        ('So no Asaas (taxas):', len(resultado['so_asaas_taxas']), yellow_fill),
        ('So no Asaas (outros):', len(resultado['so_asaas_outros']), yellow_fill),
        ('So no ADVBOX:', len(resultado['so_advbox']), blue_fill),
    ]
    for label, qtd, fill in stats:
        ws.cell(row=r, column=1, value=label).font = bold
        c = ws.cell(row=r, column=2, value=qtd)
        c.font = bold
        c.fill = fill
        r += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18

    # ============ ABA 2: CONCILIADOS ============
    ws2 = wb.create_sheet('CONCILIADOS', 1)
    headers2 = ['Data Asaas', 'Cliente', 'Valor', 'Descricao Asaas',
                'Data ADVBOX', 'Categoria ADVBOX', 'Descricao ADVBOX', 'Match']
    widths2 = [12, 25, 14, 55, 12, 25, 35, 15]
    hdr_row(ws2, 1, headers2, widths2)

    for i, c in enumerate(resultado['conciliados']):
        r = i + 2
        a = c['asaas']
        b = c['advbox']
        data_cell(ws2, r, 1, a['data'])
        data_cell(ws2, r, 2, a['nome_cliente'])
        data_cell(ws2, r, 3, a['valor'], money, green_fill)
        data_cell(ws2, r, 4, a['descricao'])
        data_cell(ws2, r, 5, b['data_venc'] or b['data_pgto'])
        data_cell(ws2, r, 6, b['categoria'])
        data_cell(ws2, r, 7, b['descricao'])
        data_cell(ws2, r, 8, c['tipo_match'])

    r = len(resultado['conciliados']) + 3
    ws2.cell(row=r, column=2, value='TOTAL CONCILIADO:').font = bold
    total_conc = sum(c['asaas']['valor'] for c in resultado['conciliados'])
    ws2.cell(row=r, column=3, value=total_conc).number_format = money
    ws2.cell(row=r, column=3).font = bold

    # ============ ABA 3: SO ASAAS (faltam no ADVBOX) ============
    ws3 = wb.create_sheet('FALTAM NO ADVBOX', 2)
    headers3 = ['Data', 'Tipo', 'Cliente', 'Valor', 'Descricao', 'Fatura', 'Tipo Asaas', 'Acao Sugerida']
    widths3 = [12, 10, 25, 14, 55, 14, 22, 25]
    hdr_row(ws3, 1, headers3, widths3)

    todos_so_asaas = resultado['so_asaas_cobrancas'] + resultado['so_asaas_taxas'] + resultado['so_asaas_outros']
    for i, a in enumerate(todos_so_asaas):
        r = i + 2
        fill = red_fill if a['eh_cobranca'] else yellow_fill
        data_cell(ws3, r, 1, a['data'])
        data_cell(ws3, r, 2, 'RECEITA' if a['tipo'] == 'receita' else 'DESPESA', fill=fill)
        data_cell(ws3, r, 3, a['nome_cliente'])
        data_cell(ws3, r, 4, a['valor'], money, fill)
        data_cell(ws3, r, 5, a['descricao'])
        data_cell(ws3, r, 6, a.get('fatura', ''))
        data_cell(ws3, r, 7, a['tipo_asaas'])
        if a['eh_cobranca']:
            acao = 'LANCAR RECEITA NO ADVBOX'
        elif a['tipo'] == 'despesa':
            acao = 'LANCAR DESPESA NO ADVBOX'
        else:
            acao = 'VERIFICAR'
        data_cell(ws3, r, 8, acao, fill=fill)

    r = len(todos_so_asaas) + 3
    ws3.cell(row=r, column=2, value='TOTAL FALTANTE:').font = bold
    ws3.cell(row=r, column=4, value=sum(a['valor'] for a in todos_so_asaas)).number_format = money
    ws3.cell(row=r, column=4).font = bold

    # ============ ABA 4: SO ADVBOX ============
    ws4 = wb.create_sheet('SO ADVBOX', 3)
    headers4 = ['Data Venc.', 'Data Pgto', 'Tipo', 'Cliente', 'Valor',
                'Categoria', 'Descricao', 'Responsavel', 'Processo']
    widths4 = [12, 12, 10, 25, 14, 25, 35, 18, 20]
    hdr_row(ws4, 1, headers4, widths4)

    for i, b in enumerate(resultado['so_advbox']):
        r = i + 2
        fill = blue_fill
        data_cell(ws4, r, 1, b['data_venc'])
        data_cell(ws4, r, 2, b['data_pgto'])
        data_cell(ws4, r, 3, 'RECEITA' if b['tipo'] == 'receita' else 'DESPESA', fill=fill)
        data_cell(ws4, r, 4, b['nome_cliente'])
        data_cell(ws4, r, 5, b['valor'], money, fill)
        data_cell(ws4, r, 6, b['categoria'])
        data_cell(ws4, r, 7, b['descricao'])
        data_cell(ws4, r, 8, b['responsavel'])
        data_cell(ws4, r, 9, b['num_processo'])

    r = len(resultado['so_advbox']) + 3
    ws4.cell(row=r, column=3, value='TOTAL:').font = bold
    ws4.cell(row=r, column=5, value=sum(b['valor'] for b in resultado['so_advbox'])).number_format = money
    ws4.cell(row=r, column=5).font = bold

    return wb


# ============================================================
# LANCAR FALTANTES NO ADVBOX (AUTOMATICO)
# ============================================================

def lancar_faltantes_advbox(resultado, confirmar=True):
    """
    Lanca no ADVBOX as transacoes que existem no Asaas mas faltam no ADVBOX.
    Somente cobrancas recebidas (receitas) e taxas (despesas).

    confirmar: se True, mostra preview antes de lancar.
    Retorna: lista de transacoes criadas.
    """
    faltantes_receitas = resultado['so_asaas_cobrancas']
    faltantes_despesas = resultado['so_asaas_taxas']

    if not faltantes_receitas and not faltantes_despesas:
        print('  Nenhuma transacao faltante para lancar.')
        return []

    print(f'\n--- LANCAMENTOS PENDENTES ---')
    print(f'  Receitas (cobrancas): {len(faltantes_receitas)}')
    for a in faltantes_receitas:
        print(f'    R$ {a["valor"]:>10,.2f} | {a["nome_cliente"]} | {a["data"]}')
    print(f'  Despesas (taxas):     {len(faltantes_despesas)}')
    for a in faltantes_despesas:
        print(f'    R$ {a["valor"]:>10,.2f} | {a["descricao"][:50]} | {a["data"]}')

    total_receitas = sum(a['valor'] for a in faltantes_receitas)
    total_despesas = sum(a['valor'] for a in faltantes_despesas)
    print(f'\n  Total receitas: R$ {total_receitas:,.2f}')
    print(f'  Total despesas: R$ {total_despesas:,.2f}')

    if confirmar:
        resp = input('\n  Lancar no ADVBOX? (s/N): ').strip().lower()
        if resp != 's':
            print('  Cancelado.')
            return []

    # IDs de referencia (buscados dinamicamente no ADVBOX do escritorio)
    cat_receita = buscar_categoria('HONOR', 'income')
    cat_despesa = buscar_categoria('TAXA', 'expense')
    if not cat_despesa:
        cat_despesa = buscar_categoria('', 'expense')
    banco_id = buscar_banco('ASAAS')
    if not banco_id:
        banco_id = buscar_banco()
    centro_custo_id = buscar_centro_custo('GERAL')
    if not centro_custo_id:
        centro_custo_id = buscar_centro_custo()

    criados = []

    # Lancar receitas
    for a in faltantes_receitas:
        cliente_id = None
        if a['nome_cliente']:
            clientes = buscar_cliente(nome=a['nome_cliente'])
            if clientes:
                cliente_id = clientes[0]['id']

        desc = f'ASAAS - {a["descricao"][:80]}' if a['descricao'] else f'ASAAS - Cobranca {a["nome_cliente"]}'
        try:
            res = criar_transacao(
                tipo='income',
                valor=a['valor'],
                data_vencimento=a['data'],
                categoria_id=cat_receita,
                banco_id=banco_id,
                centro_custo_id=centro_custo_id,
                cliente_id=cliente_id,
                descricao=desc,
                data_pagamento=a['data'],
            )
            if res and res.get('success'):
                criados.append({'tipo': 'receita', 'asaas': a, 'advbox_id': res['transactions_id']})
        except Exception as e:
            print(f'  ERRO ao lancar receita {a["nome_cliente"]}: {e}')

    # Lancar despesas (taxas)
    for a in faltantes_despesas:
        desc = f'ASAAS TAXA - {a["descricao"][:80]}' if a['descricao'] else f'ASAAS - Taxa {a["tipo_asaas"]}'
        try:
            res = criar_transacao(
                tipo='expense',
                valor=a['valor'],
                data_vencimento=a['data'],
                categoria_id=cat_despesa,
                banco_id=banco_id,
                centro_custo_id=centro_custo_id,
                descricao=desc,
                data_pagamento=a['data'],
            )
            if res and res.get('success'):
                criados.append({'tipo': 'despesa', 'asaas': a, 'advbox_id': res['transactions_id']})
        except Exception as e:
            print(f'  ERRO ao lancar despesa: {e}')

    print(f'\n  Lancados com sucesso: {len(criados)}/{len(faltantes_receitas) + len(faltantes_despesas)}')
    return criados


# ============================================================
# UPLOAD PARA GOOGLE DRIVE
# ============================================================

def upload_planilha(wb, nome_arquivo):
    """Salva planilha e faz upload pro Google Drive (pasta financeira)."""
    if not PASTA_FINANCEIRO_API:
        local_path = os.path.join(os.path.dirname(__file__), f'{nome_arquivo}.xlsx')
        wb.save(local_path)
        print(f'\n  AVISO: DRIVE_PASTA_FINANCEIRO_ID nao configurado - salvo local: {local_path}')
        return None

    print(f'\nSalvando e enviando pro Drive: {nome_arquivo}...')
    drive, _ = autenticar_google()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.read()

    media = MediaInMemoryUpload(
        file_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    file_metadata = {
        'name': nome_arquivo,
        'parents': [PASTA_FINANCEIRO_API],
        'mimeType': 'application/vnd.google-apps.spreadsheet'
    }
    http_upload = google_auth_httplib2.AuthorizedHttp(
        drive._http.credentials,
        http=httplib2.Http(timeout=300)
    )
    for tentativa in range(3):
        try:
            arquivo = drive.files().create(
                body=file_metadata, media_body=media, fields='id, webViewLink'
            ).execute(http=http_upload)
            print(f'  Arquivo criado: {arquivo.get("webViewLink")}')
            return arquivo
        except Exception as e:
            print(f'  Tentativa {tentativa+1} falhou: {e}')
            if tentativa == 2:
                raise
            import time
            time.sleep(5)
            media = MediaInMemoryUpload(
                file_bytes,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Conciliacao Financeira Asaas x ADVBOX')
    parser.add_argument('competencia', nargs='?', default=None,
                        help='Competencia MM/YYYY (ex: 04/2026). Padrao: mes atual')
    parser.add_argument('--lancar', action='store_true',
                        help='Lancar faltantes no ADVBOX automaticamente')
    parser.add_argument('--sem-upload', action='store_true',
                        help='Nao fazer upload pro Drive')
    args = parser.parse_args()

    if args.competencia:
        competencia = args.competencia
    else:
        hoje = datetime.now()
        competencia = f'{hoje.month:02d}/{hoje.year}'

    resultado = conciliar(competencia)

    print('\n4. Gerando planilha de conciliacao...')
    wb = gerar_planilha(resultado)

    nome_mes = resultado['nome_mes']
    ano = resultado['ano']
    nome_arquivo = f'CONCILIACAO {nome_mes} {ano} - ASAAS x ADVBOX'

    if not args.sem_upload:
        upload_planilha(wb, nome_arquivo)
    else:
        local_path = os.path.join(os.path.dirname(__file__), f'{nome_arquivo}.xlsx')
        wb.save(local_path)
        print(f'  Salvo local: {local_path}')

    if args.lancar:
        print('\n5. Lancando faltantes no ADVBOX...')
        criados = lancar_faltantes_advbox(resultado, confirmar=True)

    print(f'\n{"="*60}')
    print(f'  CONCILIACAO {nome_mes}/{ano} CONCLUIDA!')
    print(f'{"="*60}')


if __name__ == '__main__':
    main()
