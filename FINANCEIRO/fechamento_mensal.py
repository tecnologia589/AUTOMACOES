"""
=============================================================================
  FECHAMENTO MENSAL COMPLETO - CORBELINO ADVOGADOS ASSOCIADOS
=============================================================================

  Roda TODO o pipeline financeiro com um unico comando:

  1. Baixa extrato Asaas (receitas)
  2. Classifica comissoes (conforme regras cadastradas em config/regras_financeiras.py)
  3. Gera planilha de fechamento (aba do mes + uma aba por comissionado)
  4. Puxa transacoes ADVBOX e concilia
  5. Calcula analise financeira (% despesa, lucro, provisao)
  6. Preenche planilha de resultado anual
  7. Lanca comissoes no ADVBOX (com confirmacao)
  8. Salva tudo na pasta de fechamento no Drive
  9. Mostra contas a pagar das proximas 2 semanas

  Uso:
    py fechamento_mensal.py 03/2026 --extrato-c6 "EXTRATO.pdf" --senha 123456

  Sem extrato C6 (so Asaas + ADVBOX):
    py fechamento_mensal.py 03/2026

  Sem lancar comissoes:
    py fechamento_mensal.py 03/2026 --sem-lancar

  -----------------------------------------------------------------------------
  >>> TODO (onboarding): este pipeline le TODAS as regras de comissao e
  >>> exececoes de config/regras_financeiras.py (vazio por padrao). Cadastre la
  >>> os comissionados/percentuais/exclusoes do escritorio antes de usar em
  >>> producao. Os IDs de pasta do Drive e usuario ADVBOX vem do .env (vazios).
  -----------------------------------------------------------------------------
=============================================================================
"""
import sys, io, os, re, time, argparse
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'INTEGRACOES'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'config'))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', 'config', '.env'))

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from google_integration import autenticar_google
from googleapiclient.http import MediaInMemoryUpload
import httplib2
import google_auth_httplib2

from advbox_integration import (
    carregar_settings, listar_transacoes, buscar_cliente,
    _request as advbox_request, obter_transacao
)

# Regras de negocio configuraveis (vazias por padrao - cadastrar no onboarding)
import regras_financeiras as regras
import equipe

# ============================================================
# CONFIG (tudo via env - cliente preenche o config/.env DELE)
# ============================================================
ASAAS_API_TOKEN = os.getenv('ASAAS_API_TOKEN', '')
ASAAS_BASE_URL = os.getenv('ASAAS_BASE_URL', 'https://api.asaas.com/v3')

# IDs de planilhas/pastas do Drive - VAZIOS por padrao (cliente cria as dele).
PLANILHA_ORIGINAL_ID = os.getenv('DRIVE_PLANILHA_HISTORICO_ID', '')   # historico de meses
PLANILHA_RESULTADO_ID = os.getenv('DRIVE_PLANILHA_RESULTADO_ID', '')  # resultado anual
PASTA_FECHAMENTO_IA = os.getenv('DRIVE_PASTA_FECHAMENTO_ID', '')      # pasta raiz de fechamentos

MESES_NOME = {
    '01': 'JANEIRO', '02': 'FEVEREIRO', '03': 'MARCO', '04': 'ABRIL',
    '05': 'MAIO', '06': 'JUNHO', '07': 'JULHO', '08': 'AGOSTO',
    '09': 'SETEMBRO', '10': 'OUTUBRO', '11': 'NOVEMBRO', '12': 'DEZEMBRO'
}
MESES_ABREV = {
    '01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR',
    '05': 'MAI', '06': 'JUN', '07': 'JUL', '08': 'AGO',
    '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'
}

# Estilos
HEADER_FONT = Font(name='Montserrat', bold=True, size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT_WHITE = Font(name='Montserrat', bold=True, size=11, color='FFFFFF')
NORMAL_FONT = Font(name='Montserrat', size=10)
MONEY_FMT = '#,##0.00'
THIN_BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CREDITO_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
DEBITO_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
COMISSAO_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')


# ============================================================
# FUNCOES AUXILIARES
# ============================================================

def banner(nome_mes, ano, start_date, end_date):
    print('\n')
    print('=' * 70)
    print('   CORBELINO ADVOGADOS ASSOCIADOS - FECHAMENTO MENSAL COMPLETO')
    print(f'   {nome_mes}/{ano}')
    print('=' * 70)
    print(f'   Executado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
    print(f'   Periodo: {start_date} a {end_date}')
    print('=' * 70)
    print()


def extrair_nome_cliente(desc):
    desc = str(desc)
    if 'fatura nr.' in desc:
        desc = re.split(r'fatura nr\.\s*\d+\s*', desc)[-1]
    nome = re.split(r'\s*\[', desc)[0].strip()
    nome = re.sub(r'[_\s]*(ATUAL)$', '', nome).strip()
    nome = re.sub(r'\.\s*Nota fiscal nr\.\s*\d+', '', nome).strip()
    nome = re.sub(r'^recebida\s*-\s*ATRASADOS\s*CONTA\s*\w+\s*-\s*', '', nome).strip()
    return nome.upper()


def baixar_extrato_asaas(start_date, finish_date):
    headers = {'access_token': ASAAS_API_TOKEN}
    todas = []
    offset = 0
    while True:
        params = {'startDate': start_date, 'finishDate': finish_date, 'offset': offset, 'limit': 100}
        resp = requests.get(f'{ASAAS_BASE_URL}/financialTransactions',
                            headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        todas.extend(data['data'])
        print(f'     Asaas: {len(todas)}/{data["totalCount"]}')
        if not data['hasMore']:
            break
        offset += 100
    return todas


def extrair_c6_pdf(pdf_path, senha):
    """Extrai transacoes de um extrato bancario em PDF protegido por senha."""
    import pikepdf
    import pdfplumber

    # Decriptar
    temp_path = pdf_path.replace('.pdf', '_aberto.pdf')
    pdf = pikepdf.open(pdf_path, password=senha)
    pdf.save(temp_path)
    pdf.close()

    # Extrair texto
    saidas = []
    entradas = []
    with pdfplumber.open(temp_path) as plumber:
        for page in plumber.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                # Formato: DD/MM DD/MM Tipo Descricao R$ X.XXX,XX ou -R$ X.XXX,XX
                match = re.search(r'(\d{2}/\d{2})\s+\d{2}/\d{2}\s+(.+?)\s+(-?R\$\s*[\d.,]+)', line)
                if match:
                    data = match.group(1)
                    desc = match.group(2).strip()
                    valor_str = match.group(3).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                    valor = float(valor_str)
                    if valor < 0:
                        saidas.append((data, desc, abs(valor)))
                    else:
                        entradas.append((data, desc, valor))

    try:
        os.remove(temp_path)
    except OSError:
        pass

    return saidas, entradas


# ============================================================
# PASSO 1: FECHAMENTO ASAAS + CLASSIFICACAO
# ============================================================

def passo1_fechamento_asaas(start_date, end_date, mes, ano, drive):
    print('=' * 70)
    print('  PASSO 1: FECHAMENTO ASAAS + CLASSIFICACAO COMISSOES')
    print('=' * 70)

    # Baixar extrato Asaas
    print(f'\n  [1.1] Baixando extrato Asaas ({start_date} a {end_date})...')
    transacoes_api = baixar_extrato_asaas(start_date, end_date)
    print(f'     Total: {len(transacoes_api)} transacoes')

    # Processar e classificar comissoes a partir das regras configuraveis.
    # >>> TODO (onboarding): se COMISSOES estiver vazio, nenhum cliente sera
    # >>> classificado em comissao (parceiro == '' = escritorio).
    dados = []
    for t in transacoes_api:
        valor = t['value']
        tipo_lanc = 'Crédito' if valor > 0 else 'Débito'
        descricao = t.get('description') or ''
        parceiro = ''
        if tipo_lanc == 'Crédito' and 'Cobrança recebida' in descricao:
            parceiro = regras.classificar_comissao(descricao) or ''
        dados.append({
            'data': t.get('date', ''), 'tipo_trans': t.get('type', ''),
            'descricao': descricao, 'valor': abs(valor), 'saldo': t.get('balance', 0),
            'fatura': t.get('paymentId', ''), 'tipo_lanc': tipo_lanc, 'parceiro': parceiro
        })

    cobr_recebidas = [d for d in dados if d['tipo_lanc'] == 'Crédito' and 'Cobrança recebida' in d['descricao']]
    nome_mes = MESES_NOME[mes]
    abrev = MESES_ABREV[mes]
    ano_curto = ano[2:]

    # Classificacao
    print(f'\n  [1.2] Classificacao ({len(cobr_recebidas)} cobrancas):')
    for d in cobr_recebidas:
        nome = extrair_nome_cliente(d['descricao'])
        rotulo = regras.COMISSOES.get(d['parceiro'], {}).get('rotulo', 'ESCRITORIO') if d['parceiro'] else 'ESCRITORIO'
        print(f'     [{rotulo:^12}] R$ {d["valor"]:>10,.2f} | {nome}')

    # === GERAR PLANILHA ===
    print(f'\n  [1.3] Gerando planilha de fechamento...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- ABA 1: MES ---
    ws = wb.create_sheet(f'{nome_mes}_{ano_curto}', 0)
    ws.merge_cells('A1:H1')
    ws['A1'] = f'FECHAMENTO {nome_mes}/{ano}'
    ws['A1'].font = Font(name='Montserrat', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    creditos = [d for d in dados if d['tipo_lanc'] == 'Crédito']
    debitos = [d for d in dados if d['tipo_lanc'] == 'Débito']

    ws['A3'] = 'RESUMO DO MES'
    ws['A3'].font = Font(name='Montserrat', bold=True, size=12)
    for r, (label, val) in enumerate([
        ('Total Creditos:', sum(d['valor'] for d in creditos)),
        ('Total Debitos:', sum(d['valor'] for d in debitos)),
        ('Cobrancas Recebidas:', sum(d['valor'] for d in cobr_recebidas)),
    ], 4):
        ws.cell(row=r, column=1, value=label).font = Font(name='Montserrat', bold=True, size=10)
        ws.cell(row=r, column=2, value=val).number_format = MONEY_FMT

    # Resumo por comissionado (dinamico a partir das regras)
    linha_resumo = 4
    receita_escritorio = sum(d['valor'] for d in cobr_recebidas if not d['parceiro'])
    ws.cell(row=linha_resumo, column=4, value='ESCRITORIO:').font = Font(name='Montserrat', bold=True, size=10)
    ws.cell(row=linha_resumo, column=5, value=receita_escritorio).number_format = MONEY_FMT
    for chave, regra in regras.COMISSOES.items():
        linha_resumo += 1
        val = sum(d['valor'] for d in cobr_recebidas if d['parceiro'] == chave)
        ws.cell(row=linha_resumo, column=4, value=f"{regra.get('rotulo', chave)}:").font = Font(name='Montserrat', bold=True, size=10)
        ws.cell(row=linha_resumo, column=5, value=val).number_format = MONEY_FMT

    headers = ['Data', 'Tipo Transacao', 'Descricao', 'Valor', 'Saldo', 'Fatura',
               'Tipo', 'Comissionado', 'Comissao']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row_n = 10
    for d in dados:
        rotulo = regras.COMISSOES.get(d['parceiro'], {}).get('rotulo', '') if d['parceiro'] else ''
        for c, v in enumerate([d['data'], d['tipo_trans'], d['descricao'], d['valor'],
                                d['saldo'], d['fatura'], d['tipo_lanc'], rotulo], 1):
            cell = ws.cell(row=row_n, column=c, value=v)
            cell.font = NORMAL_FONT
            if c in (4, 5):
                cell.number_format = MONEY_FMT

        fill = CREDITO_FILL if d['tipo_lanc'] == 'Crédito' else DEBITO_FILL if d['tipo_lanc'] == 'Débito' else None
        if fill:
            for c in range(1, 8):
                ws.cell(row=row_n, column=c).fill = fill
        if d['parceiro']:
            ws.cell(row=row_n, column=8).fill = COMISSAO_FILL

        # Comissao a partir da regra cadastrada
        nome_cli = extrair_nome_cliente(d['descricao'])
        if d['parceiro'] and 'Cobrança recebida' in d['descricao']:
            regra = regras.COMISSOES.get(d['parceiro'], {})
            pct = regra.get('percentual', 0)
            if pct and regras.comissionado_recebe(d['parceiro'], nome_cli):
                ws.cell(row=row_n, column=9, value=d['valor'] * pct).number_format = MONEY_FMT

        for c in range(1, 10):
            ws.cell(row=row_n, column=c).border = THIN_BORDER
        row_n += 1

    row_t = row_n + 1
    ws.cell(row=row_t, column=3, value='TOTAL COMISSOES:').font = HEADER_FONT
    total_comissoes = sum(ws.cell(row=r, column=9).value or 0 for r in range(10, row_n))
    ws.cell(row=row_t, column=9, value=total_comissoes).number_format = MONEY_FMT
    ws.cell(row=row_t, column=9).font = HEADER_FONT

    for col, w in enumerate([12, 20, 65, 14, 14, 14, 10, 18, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # --- ABAS POR COMISSIONADO (uma por chave em COMISSOES) ---
    # >>> TODO (onboarding): sem COMISSOES cadastradas nao ha abas extras.
    totais_por_comissionado = {}
    idx_aba = 1
    for chave, regra in regras.COMISSOES.items():
        pct = regra.get('percentual', 0)
        rotulo = regra.get('rotulo', chave)
        processos = [d for d in cobr_recebidas
                     if d['parceiro'] == chave
                     and regras.comissionado_recebe(chave, extrair_nome_cliente(d['descricao']))]

        ws_c = wb.create_sheet(f'{rotulo[:20]} {abrev}.{ano_curto}', idx_aba)
        idx_aba += 1
        ws_c.merge_cells('A1:D1')
        ws_c['A1'] = f'EXTRATO {nome_mes} - COMISSAO {rotulo.upper()}'
        ws_c['A1'].font = Font(name='Montserrat', bold=True, size=13)
        ws_c['B3'] = f'Periodo de 01/{mes}/{ano} ate {end_date[8:]}/{mes}/{ano}'
        ws_c['B3'].font = Font(name='Montserrat', bold=True, size=10)

        for col, h in enumerate(['Data', 'Descricao', 'FATURAMENTO', f'COMISSAO [{pct*100:.0f}%]'], 1):
            cell = ws_c.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT_WHITE; cell.fill = HEADER_FILL; cell.border = THIN_BORDER

        c_row = 5
        for d in processos:
            ws_c.cell(row=c_row, column=1, value=d['data']).font = NORMAL_FONT
            ws_c.cell(row=c_row, column=2, value=d['descricao']).font = NORMAL_FONT
            ws_c.cell(row=c_row, column=3, value=d['valor']).number_format = MONEY_FMT
            ws_c.cell(row=c_row, column=4, value=d['valor'] * pct).number_format = MONEY_FMT
            for c in range(1, 5): ws_c.cell(row=c_row, column=c).border = THIN_BORDER
            c_row += 1

        c_row += 1
        total = sum(d['valor'] * pct for d in processos)
        ws_c.cell(row=c_row, column=2, value='TOTAIS:').font = HEADER_FONT
        ws_c.cell(row=c_row, column=3, value=sum(d['valor'] for d in processos)).number_format = MONEY_FMT
        ws_c.cell(row=c_row, column=4, value=total).number_format = MONEY_FMT
        for c in (3, 4): ws_c.cell(row=c_row, column=c).font = HEADER_FONT
        for col, w in enumerate([12, 70, 16, 24], 1):
            ws_c.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        totais_por_comissionado[chave] = {
            'rotulo': rotulo, 'total': total, 'processos': processos,
            'percentual': pct, 'advbox_customers_id': regra.get('advbox_customers_id'),
        }
        print(f'     Aba {rotulo}: {len(processos)} processos | Total: R$ {total:,.2f}')

    print(f'     Aba {nome_mes}_{ano_curto}: {row_n - 10} linhas')

    return {
        'wb': wb, 'dados': dados, 'cobr_recebidas': cobr_recebidas,
        'totais_por_comissionado': totais_por_comissionado,
    }


# ============================================================
# PASSO 2: CONCILIACAO ADVBOX
# ============================================================

def passo2_advbox(start_date, end_date, competencia):
    print('\n' + '=' * 70)
    print('  PASSO 2: DADOS ADVBOX')
    print('=' * 70)

    print('\n  [2.1] Baixando transacoes ADVBOX (por vencimento)...')
    try:
        trans = listar_transacoes({'date_due_start': start_date, 'date_due_end': end_date})

        receitas_todas = [t for t in trans if t.get('entry_type') == 'income']
        # Excluir clientes que nao contam como receita do escritorio (config)
        receitas = [t for t in receitas_todas
                    if not regras.excluir_do_faturamento(t.get('name') or '')]
        excluidos = [t for t in receitas_todas if t not in receitas]
        if excluidos:
            print(f'     Excluidos do faturamento ({len(excluidos)}):')
            for t in excluidos:
                print(f'       {t.get("name","")} | R$ {t.get("amount",0):,.2f} | {t.get("description","")[:40]}')
        despesas = [t for t in trans if t.get('entry_type') == 'expense']
        total_receitas = sum(t.get('amount', 0) for t in receitas)
        total_despesas = sum(t.get('amount', 0) for t in despesas)

        # Separar distribuicao de lucros (categorias configuraveis)
        distribuicao = sum(t.get('amount', 0) for t in despesas if regras.eh_distribuicao(t.get('category')))
        desp_operacionais = total_despesas - distribuicao

        print(f'     Receitas: {len(receitas)} = R$ {total_receitas:,.2f}')
        print(f'     Despesas: {len(despesas)} = R$ {total_despesas:,.2f}')
        print(f'     Distribuicao lucros: R$ {distribuicao:,.2f}')
        print(f'     Desp. operacionais: R$ {desp_operacionais:,.2f}')

        # Categorias
        cats = defaultdict(float)
        for t in despesas:
            cats[t.get('category') or 'SEM CATEGORIA'] += t.get('amount', 0)
        print(f'\n  [2.2] Despesas por categoria:')
        for cat, val in sorted(cats.items(), key=lambda x: -x[1]):
            print(f'     R$ {val:>10,.2f} | {cat}')

        # Pendentes de baixa
        pendentes = [t for t in despesas if not t.get('date_payment')]
        if pendentes:
            print(f'\n  [2.3] Despesas pendentes de baixa ({len(pendentes)}):')
            for t in pendentes:
                desc = (t.get('description') or '')[:45]
                nome = (t.get('name') or '')[:25]
                print(f'     ID:{t["id"]} | {t.get("date_due","?")} | R$ {t.get("amount",0):>10,.2f} | {t.get("category","")[:25]} | {nome} | {desc}')

        return {
            'total_receitas': total_receitas, 'total_despesas': total_despesas,
            'distribuicao': distribuicao, 'desp_operacionais': desp_operacionais,
            'receitas': receitas, 'despesas': despesas, 'categorias': dict(cats),
        }
    except Exception as e:
        print(f'     ERRO ADVBOX: {e}')
        return None


# ============================================================
# PASSO 3: ANALISE FINANCEIRA + PLANILHA RESULTADO
# ============================================================

def passo3_analise(advbox_data, fechamento, mes, ano, drive):
    print('\n' + '=' * 70)
    print('  PASSO 3: ANALISE FINANCEIRA')
    print('=' * 70)

    if advbox_data:
        faturamento = advbox_data['total_receitas']
        despesas_totais = advbox_data['total_despesas']
        distribuicao = advbox_data['distribuicao']
        desp_operacionais = advbox_data['desp_operacionais']
    else:
        # Fallback Asaas
        faturamento = sum(d['valor'] for d in fechamento['cobr_recebidas'])
        despesas_totais = 0
        distribuicao = 0
        desp_operacionais = 0

    lucro_liquido = faturamento - desp_operacionais
    margem = (lucro_liquido / faturamento * 100) if faturamento > 0 else 0
    perc_desp = (desp_operacionais / faturamento * 100) if faturamento > 0 else 0
    perc_distrib = (distribuicao / faturamento * 100) if faturamento > 0 else 0
    # Provisao configuravel sobre o lucro liquido (0 = desativado)
    provisao = lucro_liquido * regras.PERCENTUAL_PROVISAO_LUCRO if lucro_liquido > 0 else 0

    print(f'\n  Faturamento (ADVBOX):       R$ {faturamento:>12,.2f}')
    print(f'  Despesas Totais:            R$ {despesas_totais:>12,.2f}')
    print(f'  Distribuicao Lucros:        R$ {distribuicao:>12,.2f}')
    print(f'  Desp. Operacionais:         R$ {desp_operacionais:>12,.2f}')
    print(f'  LUCRO LIQUIDO:              R$ {lucro_liquido:>12,.2f}')
    print(f'  Margem de lucro:            {margem:.1f}%')
    print(f'  % Despesa/Receita:          {perc_desp:.1f}%')
    print(f'  % Distribuicao/Fat:         {perc_distrib:.1f}%')
    print(f'  {regras.ROTULO_PROVISAO_LUCRO}: R$ {provisao:>12,.2f}')

    # Preencher planilha resultado (so se configurada)
    if not PLANILHA_RESULTADO_ID:
        print('\n  AVISO: DRIVE_PLANILHA_RESULTADO_ID nao configurado - pulando planilha de resultado.')
    else:
        print(f'\n  Atualizando planilha de resultado...')
        try:
            content = drive.files().get_media(fileId=PLANILHA_RESULTADO_ID).execute()
            wb_res = openpyxl.load_workbook(io.BytesIO(content))
            ws_res = wb_res[ano] if ano in wb_res.sheetnames else wb_res.active
            linha = int(mes) + 3  # Jan=4, Fev=5, Mar=6...

            categorias = (advbox_data or {}).get('categorias', {})
            comissoes_total = sum(v for k, v in categorias.items()
                                  if any(p in k.upper() for p in regras.PALAVRAS_COMISSOES))
            taxas = sum(v for k, v in categorias.items()
                        if any(x in k.upper() for x in regras.PALAVRAS_TAXAS))
            sazonais = sum(v for k, v in categorias.items()
                           if any(x in k.upper() for x in regras.PALAVRAS_SAZONAIS))

            ws_res.cell(row=linha, column=1, value=MESES_NOME[mes])
            ws_res.cell(row=linha, column=2, value=round(faturamento, 2))
            ws_res.cell(row=linha, column=3, value=round(faturamento, 2))
            ws_res.cell(row=linha, column=4, value=round(faturamento, 2))
            ws_res.cell(row=linha, column=5, value=round(faturamento - comissoes_total, 2))
            ws_res.cell(row=linha, column=7, value=round(sazonais, 2))
            ws_res.cell(row=linha, column=8, value=round(comissoes_total, 2))
            ws_res.cell(row=linha, column=9, value=round(taxas, 2))
            ws_res.cell(row=linha, column=10, value=round(distribuicao, 2))
            ws_res.cell(row=linha, column=11, value=round(despesas_totais, 2))
            ws_res.cell(row=linha, column=12, value=round(lucro_liquido, 2))
            ws_res.cell(row=linha, column=13, value=round(margem, 2))

            buffer = io.BytesIO()
            wb_res.save(buffer)
            buffer.seek(0)
            media = MediaInMemoryUpload(buffer.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            drive.files().update(fileId=PLANILHA_RESULTADO_ID, media_body=media).execute()
            print(f'  Planilha resultado atualizada! Linha {linha} ({MESES_NOME[mes]})')
        except Exception as e:
            print(f'  ERRO ao atualizar resultado: {e}')

    return {
        'faturamento': faturamento, 'lucro_liquido': lucro_liquido,
        'margem': margem, 'provisao': provisao, 'perc_desp': perc_desp,
        'perc_distrib': perc_distrib,
    }


# ============================================================
# PASSO 4: LANCAR COMISSOES NO ADVBOX (COM CONFIRMACAO)
# ============================================================

def passo4_comissoes_advbox(totais_por_comissionado, mes, ano):
    print('\n' + '=' * 70)
    print('  PASSO 4: LANCAR COMISSOES NO ADVBOX')
    print('=' * 70)

    nome_mes = MESES_NOME[mes]

    if not totais_por_comissionado:
        print('\n  Nenhuma comissao a lancar (COMISSOES vazio em config/regras_financeiras.py).')
        return

    fin = regras.ADVBOX_FINANCEIRO
    if not (fin.get('banco_principal') and fin.get('categoria_comissao') and fin.get('centro_custo')):
        print('\n  AVISO: ADVBOX_FINANCEIRO incompleto em config/regras_financeiras.py.')
        print('  >>> TODO (onboarding): cadastre banco_principal, categoria_comissao e centro_custo.')
        return

    user_financeiro = equipe.id_usuario('FINANCEIRO')
    if not user_financeiro:
        print('\n  AVISO: usuario FINANCEIRO nao configurado em config/equipe.py - nao da pra lancar.')
        return

    for chave, info in totais_por_comissionado.items():
        print(f'  Comissao {info["rotulo"]}: R$ {info["total"]:,.2f}')

    resp = input('\n  Lancar comissoes no ADVBOX? (s/N): ').strip().lower()
    if resp != 's':
        print('  Cancelado.')
        return

    date_due = f'{ano}-{int(mes)+1:02d}-05' if int(mes) < 12 else f'{int(ano)+1}-01-05'

    for chave, info in totais_por_comissionado.items():
        if info['total'] <= 0:
            continue
        cust_id = info.get('advbox_customers_id')
        if not cust_id:
            print(f'  PULADO {info["rotulo"]}: sem advbox_customers_id cadastrado.')
            continue
        time.sleep(2)
        res = advbox_request('POST', '/transactions', json_data={
            'users_id': user_financeiro, 'entry_type': 'expense',
            'debit_account': fin['banco_principal'],
            'categories_id': fin['categoria_comissao'],
            'cost_centers_id': fin['centro_custo'],
            'amount': f'{info["total"]:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            'date_due': date_due,
            'customers_id': cust_id,
            'description': f'COMISSAO {info["rotulo"].upper()} {nome_mes}/{ano}',
        })
        if res and res.get('success'):
            print(f'  OK {info["rotulo"]}! ID: {res["transactions_id"]}')


# ============================================================
# PASSO 5: SALVAR NO DRIVE
# ============================================================

def passo5_upload_drive(wb_fechamento, fechamento, advbox_data, analise, nome_mes, mes, ano, drive):
    print('\n' + '=' * 70)
    print('  PASSO 5: SALVANDO NO DRIVE')
    print('=' * 70)

    if not PASTA_FECHAMENTO_IA:
        print('  AVISO: DRIVE_PASTA_FECHAMENTO_ID nao configurado - pulando upload.')
        return

    abrev = MESES_ABREV[mes]
    ano_curto = ano[2:]
    http_upload = google_auth_httplib2.AuthorizedHttp(drive._http.credentials, http=httplib2.Http(timeout=300))

    def buscar_ou_criar_pasta(nome, parent_id):
        results = drive.files().list(
            q=f"'{parent_id}' in parents and name = '{nome}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false",
            fields='files(id)'
        ).execute()
        if results.get('files'):
            return results['files'][0]['id']
        return drive.files().create(body={
            'name': nome, 'mimeType': 'application/vnd.google-apps.folder',
            'parents': [parent_id]
        }, fields='id').execute()['id']

    def upload_xlsx(wb, nome_arquivo, pasta_id):
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.read()
        media = MediaInMemoryUpload(file_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        results = drive.files().list(
            q=f"'{pasta_id}' in parents and name contains '{nome_arquivo}' and trashed = false",
            fields='files(id)'
        ).execute()
        if results.get('files'):
            drive.files().update(fileId=results['files'][0]['id'], media_body=media).execute()
            print(f'     Atualizado: {nome_arquivo}')
            return results['files'][0]['id']
        else:
            arquivo = drive.files().create(
                body={'name': nome_arquivo, 'parents': [pasta_id],
                      'mimeType': 'application/vnd.google-apps.spreadsheet'},
                media_body=media, fields='id'
            ).execute(http=http_upload)
            print(f'     Criado: {nome_arquivo}')
            return arquivo['id']

    # Estrutura de pastas
    pasta_ano = buscar_ou_criar_pasta(ano, PASTA_FECHAMENTO_IA)
    pasta_mes = buscar_ou_criar_pasta(nome_mes, pasta_ano)
    pasta_extrato = buscar_ou_criar_pasta('EXTRATO ASAAS', pasta_mes)
    pasta_planilha = buscar_ou_criar_pasta('FECHAMENTO', pasta_mes)
    pasta_conciliacao = buscar_ou_criar_pasta('ADVBOX', pasta_mes)
    pasta_comissoes = buscar_ou_criar_pasta('COMISSOES', pasta_mes)
    pasta_resultado = buscar_ou_criar_pasta('RESULTADO FINANCEIRO', pasta_mes)

    # 1. PLANILHA DE FECHAMENTO
    print('\n  [5.1] Planilha de Fechamento...')
    upload_xlsx(wb_fechamento, f'FECHAMENTO {nome_mes} {ano}', pasta_planilha)

    # 2. EXTRATO ASAAS
    print('  [5.2] Extrato Asaas...')
    wb_extrato = openpyxl.Workbook()
    ws_ext = wb_extrato.active
    ws_ext.title = f'EXTRATO ASAAS {nome_mes}'
    headers_ext = ['Data', 'Tipo', 'Descricao', 'Valor', 'Saldo', 'Fatura', 'Credito/Debito', 'Comissionado']
    for col, h in enumerate(headers_ext, 1):
        c = ws_ext.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL; c.border = THIN_BORDER
    for i, d in enumerate(fechamento['dados'], 2):
        rotulo = regras.COMISSOES.get(d['parceiro'], {}).get('rotulo', '') if d['parceiro'] else ''
        ws_ext.cell(row=i, column=1, value=d['data'])
        ws_ext.cell(row=i, column=2, value=d['tipo_trans'])
        ws_ext.cell(row=i, column=3, value=d['descricao'])
        ws_ext.cell(row=i, column=4, value=d['valor']).number_format = MONEY_FMT
        ws_ext.cell(row=i, column=5, value=d['saldo']).number_format = MONEY_FMT
        ws_ext.cell(row=i, column=6, value=d['fatura'])
        ws_ext.cell(row=i, column=7, value=d['tipo_lanc'])
        ws_ext.cell(row=i, column=8, value=rotulo)
    ws_ext.column_dimensions['C'].width = 65
    ws_ext.column_dimensions['D'].width = 14
    upload_xlsx(wb_extrato, f'EXTRATO ASAAS {nome_mes} {ano}', pasta_extrato)

    # 3. CONCILIACAO ADVBOX
    print('  [5.3] Conciliacao ADVBOX...')
    if advbox_data:
        wb_conc = openpyxl.Workbook()
        wb_conc.remove(wb_conc.active)
        ws_rec = wb_conc.create_sheet('RECEITAS ADVBOX', 0)
        for col, h in enumerate(['Data', 'Cliente', 'Valor', 'Categoria', 'Descricao', 'Status'], 1):
            c = ws_rec.cell(row=1, column=col, value=h)
            c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL; c.border = THIN_BORDER
        for i, t in enumerate(sorted(advbox_data['receitas'], key=lambda x: x.get('date_due', '')), 2):
            ws_rec.cell(row=i, column=1, value=t.get('date_due', ''))
            ws_rec.cell(row=i, column=2, value=t.get('name', ''))
            ws_rec.cell(row=i, column=3, value=t.get('amount', 0)).number_format = MONEY_FMT
            ws_rec.cell(row=i, column=4, value=t.get('category', ''))
            ws_rec.cell(row=i, column=5, value=t.get('description', ''))
            ws_rec.cell(row=i, column=6, value='PAGO' if t.get('date_payment') else 'PENDENTE')
        ws_rec.column_dimensions['B'].width = 35
        ws_rec.column_dimensions['D'].width = 25
        ws_rec.column_dimensions['E'].width = 50
        ws_desp = wb_conc.create_sheet('DESPESAS ADVBOX', 1)
        for col, h in enumerate(['Data', 'Cliente', 'Valor', 'Categoria', 'Descricao', 'Status'], 1):
            c = ws_desp.cell(row=1, column=col, value=h)
            c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL; c.border = THIN_BORDER
        for i, t in enumerate(sorted(advbox_data['despesas'], key=lambda x: x.get('date_due', '')), 2):
            ws_desp.cell(row=i, column=1, value=t.get('date_due', ''))
            ws_desp.cell(row=i, column=2, value=t.get('name', ''))
            ws_desp.cell(row=i, column=3, value=t.get('amount', 0)).number_format = MONEY_FMT
            ws_desp.cell(row=i, column=4, value=t.get('category', ''))
            ws_desp.cell(row=i, column=5, value=t.get('description', ''))
            ws_desp.cell(row=i, column=6, value='PAGO' if t.get('date_payment') else 'PENDENTE')
        ws_desp.column_dimensions['B'].width = 35
        ws_desp.column_dimensions['D'].width = 25
        ws_desp.column_dimensions['E'].width = 50
        upload_xlsx(wb_conc, f'CONCILIACAO ADVBOX {nome_mes} {ano}', pasta_conciliacao)
    else:
        print('     ADVBOX indisponivel - pulando')

    # 4. COMISSOES (uma planilha por comissionado cadastrado)
    print('  [5.4] Comissoes...')
    for chave, info in fechamento['totais_por_comissionado'].items():
        wb_c = openpyxl.Workbook()
        ws_c = wb_c.active
        rotulo = info['rotulo']
        pct = info['percentual']
        ws_c.title = f'{rotulo[:20]} {abrev}.{ano_curto}'
        ws_c.merge_cells('A1:D1')
        ws_c['A1'] = f'COMISSAO {rotulo.upper()} - {nome_mes}/{ano}'
        ws_c['A1'].font = Font(name='Montserrat', bold=True, size=13)
        for col, h in enumerate(['Data', 'Descricao', 'FATURAMENTO', f'COMISSAO {pct*100:.0f}%'], 1):
            c = ws_c.cell(row=3, column=col, value=h)
            c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL; c.border = THIN_BORDER
        cr = 4
        for d in info['processos']:
            ws_c.cell(row=cr, column=1, value=d['data'])
            ws_c.cell(row=cr, column=2, value=d['descricao'])
            ws_c.cell(row=cr, column=3, value=d['valor']).number_format = MONEY_FMT
            ws_c.cell(row=cr, column=4, value=d['valor'] * pct).number_format = MONEY_FMT
            for c in range(1, 5): ws_c.cell(row=cr, column=c).border = THIN_BORDER
            cr += 1
        cr += 1
        ws_c.cell(row=cr, column=2, value='TOTAL COMISSAO:').font = Font(name='Montserrat', bold=True, size=12)
        ws_c.cell(row=cr, column=4, value=info['total']).number_format = MONEY_FMT
        ws_c.cell(row=cr, column=4).font = Font(name='Montserrat', bold=True, size=12)
        ws_c.column_dimensions['B'].width = 70
        ws_c.column_dimensions['C'].width = 16
        ws_c.column_dimensions['D'].width = 20
        upload_xlsx(wb_c, f'COMISSAO {rotulo.upper()} {nome_mes} {ano}', pasta_comissoes)
    if not fechamento['totais_por_comissionado']:
        print('     Sem comissoes cadastradas - pulando')

    # 5. RESULTADO FINANCEIRO
    print('  [5.5] Resultado financeiro...')
    wb_res = openpyxl.Workbook()
    ws_res = wb_res.active
    ws_res.title = f'RESULTADO {nome_mes}'
    ws_res['A1'] = f'RESULTADO FINANCEIRO {nome_mes}/{ano}'
    ws_res['A1'].font = Font(name='Montserrat', bold=True, size=14)
    dados_res = [
        ('Faturamento (ADVBOX):', analise.get('faturamento', 0)),
        ('Despesas Operacionais:', analise.get('faturamento', 0) - analise.get('lucro_liquido', 0)),
        ('Lucro Liquido:', analise.get('lucro_liquido', 0)),
        ('', ''),
        ('Margem de Lucro:', f"{analise.get('margem', 0):.1f}%"),
        ('% Despesa/Receita:', f"{analise.get('perc_desp', 0):.1f}%"),
        ('% Distribuicao/Faturamento:', f"{analise.get('perc_distrib', 0):.1f}%"),
        ('', ''),
        (f'{regras.ROTULO_PROVISAO_LUCRO}:', analise.get('provisao', 0)),
        ('', ''),
    ]
    for chave, info in fechamento['totais_por_comissionado'].items():
        dados_res.append((f'Comissao {info["rotulo"]}:', info['total']))
    for i, (label, val) in enumerate(dados_res, 3):
        ws_res.cell(row=i, column=1, value=label).font = Font(name='Montserrat', bold=True, size=11)
        c = ws_res.cell(row=i, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = MONEY_FMT
        c.font = Font(name='Montserrat', size=11)
    ws_res.column_dimensions['A'].width = 35
    ws_res.column_dimensions['B'].width = 20
    upload_xlsx(wb_res, f'RESULTADO {nome_mes} {ano}', pasta_resultado)

    print(f'\n  Pasta completa: https://drive.google.com/drive/folders/{pasta_mes}')


# ============================================================
# PASSO 6: CONTAS A PAGAR
# ============================================================

def passo6_contas_pagar():
    print('\n' + '=' * 70)
    print('  PASSO 6: CONTAS A PAGAR (PROXIMAS 2 SEMANAS)')
    print('=' * 70)

    hoje = datetime.now()
    fim = hoje + timedelta(days=14)

    try:
        filtros = {'date_due_start': hoje.strftime('%Y-%m-%d'), 'date_due_end': fim.strftime('%Y-%m-%d')}
        trans = listar_transacoes(filtros)

        despesas = [t for t in trans if t.get('entry_type') == 'expense' and not t.get('date_payment')]
        receitas = [t for t in trans if t.get('entry_type') == 'income' and not t.get('date_payment')]

        if despesas:
            print(f'\n  DESPESAS A PAGAR ({len(despesas)}):')
            for t in sorted(despesas, key=lambda x: x.get('date_due', '')):
                nome = (t.get('name') or '')[:25]
                cat = (t.get('category') or '')[:25]
                desc = (t.get('description') or '')[:40]
                print(f'  {t.get("date_due","?")} | R$ {t.get("amount",0):>10,.2f} | {cat} | {nome} | {desc}')
            total = sum(t.get('amount', 0) for t in despesas)
            print(f'\n  TOTAL A PAGAR: R$ {total:,.2f}')

        if receitas:
            total_rec = sum(t.get('amount', 0) for t in receitas)
            print(f'\n  RECEITAS A RECEBER ({len(receitas)}): R$ {total_rec:,.2f}')
    except Exception as e:
        print(f'  ERRO: {e}')


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Fechamento Mensal Completo - Corbelino Advogados Associados')
    parser.add_argument('competencia', help='Competencia MM/YYYY (ex: 03/2026)')
    parser.add_argument('--extrato-c6', help='PDF do extrato bancario (opcional)')
    parser.add_argument('--senha', help='Senha do PDF do extrato')
    parser.add_argument('--sem-lancar', action='store_true', help='Nao lancar comissoes no ADVBOX')
    args = parser.parse_args()

    mes, ano = args.competencia.split('/')
    start_date = f'{ano}-{mes}-01'
    if int(mes) == 12:
        end_dt = datetime(int(ano) + 1, 1, 1) - timedelta(days=1)
    else:
        end_dt = datetime(int(ano), int(mes) + 1, 1) - timedelta(days=1)
    end_date = end_dt.strftime('%Y-%m-%d')
    nome_mes = MESES_NOME[mes]

    banner(nome_mes, ano, start_date, end_date)

    # Autenticar
    print('  Autenticando Google Drive...')
    drive, _ = autenticar_google()

    # PASSO 1
    fechamento = passo1_fechamento_asaas(start_date, end_date, mes, ano, drive)

    # PASSO 2
    advbox_data = passo2_advbox(start_date, end_date, args.competencia)

    # PASSO 3
    analise = passo3_analise(advbox_data, fechamento, mes, ano, drive)

    # PASSO 4
    if not args.sem_lancar:
        passo4_comissoes_advbox(fechamento['totais_por_comissionado'], mes, ano)

    # PASSO 5
    passo5_upload_drive(fechamento['wb'], fechamento, advbox_data, analise, nome_mes, mes, ano, drive)

    # PASSO 6
    passo6_contas_pagar()

    # RESUMO FINAL
    print('\n' + '=' * 70)
    print('  FECHAMENTO CONCLUIDO!')
    print('=' * 70)
    print(f'  Mes: {nome_mes}/{ano}')
    print(f'  Faturamento:    R$ {analise["faturamento"]:>12,.2f}')
    print(f'  Lucro Liquido:  R$ {analise["lucro_liquido"]:>12,.2f}')
    print(f'  Margem:         {analise["margem"]:.1f}%')
    print(f'  {regras.ROTULO_PROVISAO_LUCRO}: R$ {analise["provisao"]:>12,.2f}')
    for chave, info in fechamento['totais_por_comissionado'].items():
        print(f'  Comissao {info["rotulo"]:<12}: R$ {info["total"]:>12,.2f}')
    print('=' * 70)


if __name__ == '__main__':
    main()
